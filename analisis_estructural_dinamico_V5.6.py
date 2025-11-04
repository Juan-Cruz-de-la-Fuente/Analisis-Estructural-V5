# -----------------------------------------------------------------
# 1. IMPORTACIÓN DE LIBRERÍAS
# -----------------------------------------------------------------
import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patches import Circle
import math
from datetime import datetime
import io
import plotly.graph_objects as go
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import base64
import tempfile
import os
from scipy.linalg import eig

# Importar openpyxl si está disponible (para Excel)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# -----------------------------------------------------------------
# 2. CONFIGURACIÓN DE PÁGINA Y CSS
# -----------------------------------------------------------------

# Configuración de la página con tema moderno
st.set_page_config(
    page_title="Análisis Estructural - Método de Matrices",
    page_icon="⚫",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS personalizado (Tomado de V4.7 para mejor UI)
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    :root {
        --primary-black: #000000;
        --primary-white: #ffffff;
        --gray-100: #f8f9fa;
        --gray-200: #e9ecef;
        --gray-300: #dee2e6;
        --gray-400: #ced4da;
        --gray-500: #adb5bd;
        --gray-600: #6c757d;
        --gray-700: #495057;
        --gray-800: #343a40;
        --gray-900: #212529;
        --blue-500: #495057;
        --blue-600: #343a40;
        --green-500: #28a745;
        --green-600: #218838;
    }
    
    .css-1d391kg {
        display: none;
    }
    
    .show-sidebar .css-1d391kg {
        display: block !important;
        background: linear-gradient(135deg, var(--gray-100) 0%, var(--gray-200) 100%);
        border-right: 2px solid var(--gray-300);
        box-shadow: 2px 0 10px rgba(0,0,0,0.1);
    }
    
    .main {
        font-family: 'Inter', sans-serif;
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
        color: var(--primary-black);
        min-height: 100vh;
    }
    
    .landing-container {
        background: linear-gradient(135deg, #2d3748 0%, #4a5568 100%);
        min-height: 100vh;
        display: flex;
        align-items: center;
        justify-content: center;
        margin: -1rem;
        padding: 2rem;
    }
    
    .mode-card {
        background: rgba(255, 255, 255, 0.25);
        backdrop-filter: blur(10px);
        border-radius: 20px;
        border: 1px solid rgba(255, 255, 255, 0.18);
        padding: 2rem;
        margin: 1rem;
        transition: all 0.3s ease;
        box-shadow: 0 8px 32px rgba(31, 38, 135, 0.37);
    }
    
    .mode-card:hover {
        transform: translateY(-10px);
        box-shadow: 0 15px 40px rgba(31, 38, 135, 0.5);
        background: rgba(255, 255, 255, 0.35);
    }
    
    .progress-bar {
        background: var(--primary-white);
        border-bottom: 3px solid var(--gray-300);
        padding: 1.5rem 0;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        position: sticky;
        top: 0;
        z-index: 100;
    }
    
    .progress-steps {
        display: flex;
        justify-content: center;
        align-items: center;
        gap: 1rem;
        margin-bottom: 1rem;
    }
    
    .progress-step {
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    
    .step-circle {
        width: 40px;
        height: 40px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 600;
        font-size: 14px;
        transition: all 0.3s ease;
    }
    
    .step-circle.completed {
        background: var(--green-500);
        color: white;
        box-shadow: 0 4px 12px rgba(16, 185, 129, 0.3);
    }
    
    .step-circle.current {
        background: var(--blue-500);
        color: white;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3);
        animation: pulse 2s infinite;
    }
    
    .step-circle.pending {
        background: var(--gray-300);
        color: var(--gray-600);
    }
    
    .step-line {
        width: 60px;
        height: 3px;
        background: var(--gray-300);
        transition: all 0.3s ease;
    }
    
    .step-line.completed {
        background: var(--green-500);
    }
    
    @keyframes pulse {
        0% { box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3); }
        50% { box-shadow: 0 4px 20px rgba(59, 130, 246, 0.6); }
        100% { box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3); }
    }
    
    h1, h2, h3 {
        font-family: 'Inter', sans-serif;
        font-weight: 700;
        color: var(--primary-black);
        letter-spacing: -0.02em;
    }
    
    h1 {
        font-size: 3rem;
        margin-bottom: 2rem;
        text-align: center;
        background: linear-gradient(135deg, #2d3748 0%, #4a5568 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    
    h2 {
        font-size: 2rem;
        margin-bottom: 1.5rem;
        border-bottom: 3px solid var(--gray-700);
        padding-bottom: 0.5rem;
        display: inline-block;
    }
    
    h3 {
        font-size: 1.5rem;
        margin-bottom: 1rem;
        color: var(--gray-800);
    }
    
    .stButton > button {
        font-family: 'Inter', sans-serif;
        font-weight: 600;
        border-radius: 12px;
        border: none;
        background: linear-gradient(135deg, var(--blue-500) 0%, var(--blue-600) 100%);
        color: var(--primary-white);
        transition: all 0.3s ease;
        padding: 0.75rem 2rem;
        font-size: 1rem;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.3);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(59, 130, 246, 0.4);
        background: linear-gradient(135deg, var(--blue-600) 0%, var(--blue-500) 100%);
    }
    
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div > select {
        font-family: 'Inter', sans-serif;
        border: 2px solid var(--gray-300);
        border-radius: 10px;
        background-color: var(--primary-white);
        color: var(--primary-black);
        padding: 0.75rem;
        transition: all 0.3s ease;
    }
    
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus,
    .stSelectbox > div > div > select:focus {
        border-color: var(--blue-500);
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
    }
    
    .metric-container {
        background: linear-gradient(135deg, var(--primary-white) 0%, var(--gray-100) 100%);
        padding: 1.5rem;
        border-radius: 15px;
        border: 1px solid var(--gray-200);
        margin: 0.5rem 0;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        transition: all 0.3s ease;
    }
    
    .metric-container:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.15);
    }
    
    .dataframe {
        font-family: 'Inter', sans-serif;
        border: none;
        border-radius: 15px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    
    .streamlit-expanderHeader {
        font-family: 'Inter', sans-serif;
        font-weight: 600;
        background: linear-gradient(135deg, var(--gray-100) 0%, var(--gray-200) 100%);
        border: 1px solid var(--gray-300);
        border-radius: 12px;
        transition: all 0.3s ease;
    }
    
    .streamlit-expanderHeader:hover {
        background: linear-gradient(135deg, var(--gray-200) 0%, var(--gray-300) 100%);
    }
    
    .stInfo {
        background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%);
        border-left: 4px solid var(--blue-500);
        color: var(--primary-black);
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.1);
    }
    
    .stSuccess {
        background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%);
        border-left: 4px solid var(--green-500);
        color: var(--primary-black);
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(16, 185, 129, 0.1);
    }
    
    .stWarning {
        background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
        border-left: 4px solid #f59e0b;
        color: var(--primary-black);
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(245, 158, 11, 0.1);
    }
    
    .stError {
        background: linear-gradient(135deg, #fee2e2 0%, #fecaca 100%);
        border-left: 4px solid #ef4444;
        color: var(--primary-black);
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(239, 68, 68, 0.1);
    }
    
    .footer-section {
        background: linear-gradient(135deg, var(--gray-900) 0%, var(--gray-800) 100%);
        color: var(--primary-white);
        padding: 3rem 0;
        margin-top: 4rem;
        border-radius: 20px 20px 0 0;
    }
    
    .footer-content {
        text-align: center;
        max-width: 800px;
        margin: 0 auto;
        padding: 0 2rem;
    }
    
    .footer-title {
        font-size: 2rem;
        font-weight: 700;
        margin-bottom: 1rem;
        background: linear-gradient(135deg, #4a5568 0%, #2d3748 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    
    .footer-survey {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        border-radius: 15px;
        padding: 2rem;
        margin: 2rem 0;
        border: 1px solid rgba(255, 255, 255, 0.2);
    }
    
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(30px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .fade-in-up {
        animation: fadeInUp 0.6s ease-out;
    }
</style>
""", unsafe_allow_html=True)


# -----------------------------------------------------------------
# 3. DEFINICIÓN DE CONSTANTES
# -----------------------------------------------------------------

# Base de datos de materiales aeroespaciales
MATERIALES_AEROESPACIALES = {
    "Aluminio 6061-T6": {
        "modulo_young": 68.9e9,
        "densidad": 2700,
        "descripcion": "Aleación de aluminio estructural común"
    },
    "Aluminio 7075-T6": {
        "modulo_young": 71.7e9,
        "densidad": 2810,
        "descripcion": "Aleación de aluminio de alta resistencia"
    },
    "Aluminio 2024-T3": {
        "modulo_young": 73.1e9,
        "densidad": 2780,
        "descripcion": "Aleación de aluminio para fuselajes"
    },
    "Titanio Ti-6Al-4V": {
        "modulo_young": 113.8e9,
        "densidad": 4430,
        "descripcion": "Aleación de titanio aeroespacial"
    },
    "Acero 4130": {
        "modulo_young": 205e9,
        "densidad": 7850,
        "descripcion": "Acero aleado para estructuras"
    },
    "Fibra de Carbono T300": {
        "modulo_young": 230e9,
        "densidad": 1760,
        "descripcion": "Compuesto de fibra de carbono"
    },
    "Magnesio AZ31B": {
        "modulo_young": 45e9,
        "densidad": 1770,
        "descripcion": "Aleación de magnesio ligera"
    }
}


# -----------------------------------------------------------------
# 4. DEFINICIÓN DE FUNCIONES
# -----------------------------------------------------------------

# --- Funciones de Utilidad y Navegación ---

def formatear_unidades(valor, tipo="presion"):
    """Formatear valores con prefijos apropiados"""
    abs_valor = abs(valor)

    if tipo == "presion":
        if abs_valor == 0:
            return "0 Pa"
        elif abs_valor < 10:
            return f"{valor:.3f} Pa"
        elif abs_valor < 1000:
            return f"{valor:.1f} Pa"
        elif abs_valor < 1e6:
            return f"{valor/1e3:.3f} kPa"
        elif abs_valor < 1e9:
            return f"{valor/1e6:.3f} MPa"
        else:
            return f"{valor/1e9:.3f} GPa"

    elif tipo == "fuerza":
        if abs_valor == 0:
            return "0 N"
        elif abs_valor < 1000:
            return f"{valor:.3f} N"
        elif abs_valor < 1e6:
            return f"{valor/1e3:.3f} kN"
        else:
            return f"{valor/1e6:.3f} MN"

    elif tipo == "desplazamiento":
        if abs_valor == 0:
            return "0 m"
        elif abs_valor < 1e-6:
            return f"{valor*1e9:.3f} nm"
        elif abs_valor < 1e-3:
            return f"{valor*1e6:.3f} μm"
        elif abs_valor < 1:
            return f"{valor*1e3:.3f} mm"
        else:
            return f"{valor:.6f} m"

    elif tipo == "rigidez":
        if abs_valor == 0:
            return "0 N/m"
        elif abs_valor < 1e-6:
            return f"{valor*1e9:.3f} nN/m"
        elif abs_valor < 1e-3:
            return f"{valor*1e6:.3f} μN/m"
        elif abs_valor < 1:
            return f"{valor*1e3:.3f} mN/m"
        elif abs_valor < 1e3:
            return f"{valor:.3f} N/m"
        elif abs_valor < 1e6:
            return f"{valor/1e3:.3f} kN/m"
        elif abs_valor < 1e9:
            return f"{valor/1e6:.3f} MN/m"
        else:
            return f"{valor/1e9:.3f} GN/m"

    return f"{valor:.6e}"

def reset_app():
    """Reiniciar la aplicación"""
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

def set_tipo_analisis(tipo):
    """Establecer el tipo de análisis (estático o dinámico)"""
    st.session_state.tipo_analisis = tipo
    st.session_state.step = 1
    st.rerun()

def set_tipo_elemento(tipo):
    """Establecer el tipo de elemento"""
    st.session_state.tipo_elemento = tipo
    st.session_state.step = 2  # Avanzar al siguiente paso (información de usuario)
    st.rerun()

def set_modo(modo):
    """Establecer el modo de análisis"""
    st.session_state.modo = modo
    st.session_state.step = 4  # Ajustar el paso según la nueva estructura
    st.rerun()

def next_step():
    """Avanzar al siguiente paso"""
    st.session_state.step += 1
    st.rerun()

def prev_step():
    """Retroceder al paso anterior"""
    if st.session_state.step > 1:
        st.session_state.step -= 1
        st.rerun()

# --- Funciones de Cálculo de Matrices (Estáticas y Dinámicas) ---

def generar_matriz_masa_barra(rho, A, L):
    """Generar matriz de masa para barra (4x4) - consitente"""
    m = rho * A * L
    matriz_masa = (m / 6) * np.array([
        [2, 0, 1, 0],
        [0, 0, 0, 0],
        [1, 0, 2, 0],
        [0, 0, 0, 0]
    ])
    return matriz_masa

def generar_matriz_masa_viga(rho, A, L):
    """Generar matriz de masa para viga pura (4x4) - consistente"""
    m = rho * A * L
    matriz_masa = (m / 420) * np.array([
        [156,  22*L,  54, -13*L],
        [22*L, 4*L**2, 13*L, -3*L**2],
        [54,  13*L,  156, -22*L],
        [-13*L, -3*L**2, -22*L, 4*L**2]
    ])
    return matriz_masa

def generar_matriz_masa_viga_portico(rho, A, L):
    """Generar matriz de masa para viga pórtico (6x6) - consistente"""
    m = rho * A * L
    # Matriz de masa local de viga pórtico (consistente)
    m_local = np.array([
        [m/3,  0,           0,        m/6,  0,           0],
        [0,    13*m/35,     11*m*L/210, 0,   9*m/70,     -13*m*L/420],
        [0,    11*m*L/210,  m*L**2/105, 0,  13*m*L/420, -m*L**2/140],
        [m/6,  0,           0,        m/3,  0,           0],
        [0,    9*m/70,      13*m*L/420, 0,   13*m/35,   -11*m*L/210],
        [0,   -13*m*L/420, -m*L**2/140, 0,  -11*m*L/210, m*L**2/105]
    ])
    return m_local

def generar_matriz_rigidez_barra(E, A, L, beta):
    """Generar matriz de rigidez para barra (4x4)"""
    c = math.cos(beta)
    s = math.sin(beta)
    factor = (E * A) / L
    
    matriz_global = factor * np.array([
        [c**2,      c*s,       -c**2,     -c*s],
        [c*s,       s**2,      -c*s,      -s**2],
        [-c**2,     -c*s,      c**2,      c*s],
        [-c*s,      -s**2,     c*s,       s**2]
    ])

    k_local = factor * np.array([
        [1, 0, -1, 0],
        [0, 0,  0, 0],
        [-1,0,  1, 0],
        [0, 0,  0, 0]
    ])

    return matriz_global, k_local

def generar_matriz_rigidez_viga(E, I, L):
    """Generar matriz de rigidez para viga pura (4x4)"""
    matriz = np.array([
        [12*E*I/L**3,  6*E*I/L**2,   -12*E*I/L**3,  6*E*I/L**2],
        [6*E*I/L**2,   4*E*I/L,      -6*E*I/L**2,   2*E*I/L],
        [-12*E*I/L**3, -6*E*I/L**2,  12*E*I/L**3,   -6*E*I/L**2],
        [6*E*I/L**2,   2*E*I/L,      -6*E*I/L**2,   4*E*I/L]
    ])
    return matriz, matriz

def generar_matriz_rigidez_viga_portico(E, A, I, L, beta):
    """Generar matriz de rigidez para viga pórtico (6x6)"""
    c = math.cos(beta)
    s = math.sin(beta)
    
    k_local = np.array([
        [E*A/L,     0,           0,        -E*A/L,    0,           0],
        [0,         12*E*I/L**3, 6*E*I/L**2, 0,       -12*E*I/L**3, 6*E*I/L**2],
        [0,         6*E*I/L**2,  4*E*I/L,   0,        -6*E*I/L**2,  2*E*I/L],
        [-E*A/L,    0,           0,         E*A/L,    0,           0],
        [0,         -12*E*I/L**3, -6*E*I/L**2, 0,      12*E*I/L**3, -6*E*I/L**2],
        [0,         6*E*I/L**2,  2*E*I/L,   0,        -6*E*I/L**2,  4*E*I/L]
    ])
    
    T = np.array([
        [c,  s,  0,  0,  0,  0],
        [-s, c,  0,  0,  0,  0],
        [0,  0,  1,  0,  0,  0],
        [0,  0,  0,  c,  s,  0],
        [0,  0,  0, -s,  c,  0],
        [0,  0,  0,  0,  0,  1]
    ])
    
    k_global = T.T @ k_local @ T
    return k_global, k_local

def generar_matriz_transformacion_viga_portico(beta):
    """Generar matriz de transformación de coordenadas para viga pórtico"""
    c = math.cos(beta)
    s = math.sin(beta)
    
    T = np.array([
        [c,  s,  0,  0,  0,  0],
        [-s, c,  0,  0,  0,  0],
        [0,  0,  1,  0,  0,  0],
        [0,  0,  0,  c,  s,  0],
        [0,  0,  0, -s,  c,  0],
        [0,  0,  0,  0,  0,  1]
    ])
    return T

# --- Funciones de Geometría y Ensamblaje ---

def calcular_grados_libertad_globales(nodo_id):
    """Calcular grados de libertad globales para un nodo (USADO POR MODO INTERACTIVO)"""
    if st.session_state.tipo_elemento == "barra":
        gl_por_nodo = 2
        return [(nodo_id - 1) * gl_por_nodo + 1, (nodo_id - 1) * gl_por_nodo + 2]
    elif st.session_state.tipo_elemento == "viga":
        gl_por_nodo = 2
        return [(nodo_id - 1) * gl_por_nodo + 1, (nodo_id - 1) * gl_por_nodo + 2]
    elif st.session_state.tipo_elemento == "viga_portico":
        gl_por_nodo = 3
        return [(nodo_id - 1) * gl_por_nodo + 1, (nodo_id - 1) * gl_por_nodo + 2, (nodo_id - 1) * gl_por_nodo + 3]
    else:
        gl_por_nodo = 2
        return [(nodo_id - 1) * gl_por_nodo + 1, (nodo_id - 1) * gl_por_nodo + 2]

def calcular_longitud_elemento(nodo_inicio, nodo_fin):
    """Calcular la longitud del elemento"""
    dx = nodo_fin['x'] - nodo_inicio['x']
    dy = nodo_fin['y'] - nodo_inicio['y']
    return math.sqrt(dx**2 + dy**2)

def calcular_angulo_beta(nodo_inicio, nodo_fin):
    """Calcular el ángulo β entre la horizontal y la barra"""
    dx = nodo_fin['x'] - nodo_inicio['x']
    dy = nodo_fin['y'] - nodo_inicio['y']
    return math.atan2(dy, dx)

def calcular_area_seccion(tipo_seccion, parametros):
    """Calcular el área de la sección según su tipo"""
    if tipo_seccion == "circular_solida":
        radio = parametros.get("radio", 0)
        return math.pi * radio**2
    elif tipo_seccion == "circular_hueca":
        radio_ext = parametros.get("radio_ext", 0)
        radio_int = parametros.get("radio_int", 0)
        return math.pi * (radio_ext**2 - radio_int**2)
    elif tipo_seccion == "rectangular":
        lado1 = parametros.get("lado1", 0)
        lado2 = parametros.get("lado2", 0)
        return lado1 * lado2
    elif tipo_seccion == "cuadrada":
        lado = parametros.get("lado", 0)
        return lado**2
    else:
        return parametros.get("area", 0.01)

def calcular_momento_inercia(tipo_seccion, parametros):
    """Calcular el momento de inercia según el tipo de sección"""
    if tipo_seccion == "circular_solida":
        radio = parametros.get("radio", 0)
        return (math.pi * radio**4) / 4
    elif tipo_seccion == "circular_hueca":
        radio_ext = parametros.get("radio_ext", 0)
        radio_int = parametros.get("radio_int", 0)
        return (math.pi * (radio_ext**4 - radio_int**4)) / 4
    elif tipo_seccion == "rectangular":
        lado1 = parametros.get("lado1", 0)  # base
        lado2 = parametros.get("lado2", 0)  # altura
        return (lado1 * lado2**3) / 12
    elif tipo_seccion == "cuadrada":
        lado = parametros.get("lado", 0)
        return (lado**4) / 12
    else:
        return parametros.get("inercia", 1e-6)

def calcular_y_asignar_grados_libertad():
    """Calcula los grados de libertad globales y la información de GL para todos los nodos y elementos."""
    st.session_state.grados_libertad_info = []
    # NO limpiar matrices de elementos aquí, se calculan en el paso 8
    # st.session_state.matrices_elementos = {} 

    # Generar información de GL para cada nodo
    gl_counter = 1
    gl_info_dict = {} # Para mapear GL number a info
    all_gl_info = [] # Lista temporal para todos los GL, incluidos los fijos
    
    gl_por_nodo_map = {
        "barra": 2,
        "viga": 2,
        "viga_portico": 3
    }
    gl_por_nodo = gl_por_nodo_map.get(st.session_state.tipo_elemento, 3)

    for nodo in st.session_state.nodos:
        nodo_id = nodo['id']
        nodo_tipo = nodo.get('tipo', 'libre') # Asumir libre si no está clasificado
        
        gl_indices = []
        
        if st.session_state.tipo_elemento == "barra":
            gl_nombres = ['X', 'Y']
        elif st.session_state.tipo_elemento == "viga":
            gl_nombres = ['Y', 'Theta']
        else: # viga_portico
            gl_nombres = ['X', 'Y', 'Theta']
        
        for i in range(gl_por_nodo):
            gl_nombre = gl_nombres[i]
            if nodo_tipo == "libre":
                gl_num = gl_counter
                gl_indices.append(gl_num)
                all_gl_info.append({
                    'numero': gl_num, 
                    'nodo': nodo_id, 
                    'direccion': gl_nombre, 
                    'fuerza_conocida': False, 
                    'valor_fuerza': 0.0, 
                    'desplazamiento_conocido': False, 
                    'valor_desplazamiento': 0.0
                })
                gl_counter += 1
            else: # Fijo
                gl_indices.append(None) # Marcador para GL fijo
                all_gl_info.append({
                    'numero': None, 
                    'nodo': nodo_id, 
                    'direccion': gl_nombre, 
                    'fuerza_conocida': False, 
                    'valor_fuerza': 0.0, 
                    'desplazamiento_conocido': True, 
                    'valor_desplazamiento': 0.0
                })
        
        nodo['grados_libertad_globales'] = [gl for gl in gl_indices if gl is not None]
        gl_info_dict[nodo_id] = gl_indices # Guardar para uso en elementos

    # Asignar GL globales a los elementos
    for elemento in st.session_state.elementos:
        nodo_inicio_id = elemento['nodo_inicio']
        nodo_fin_id = elemento['nodo_fin']
        
        # Obtener los GL de los nodos (pueden contener 'None')
        gl_inicio_raw = gl_info_dict.get(nodo_inicio_id, [None]*gl_por_nodo)
        gl_fin_raw = gl_info_dict.get(nodo_fin_id, [None]*gl_por_nodo)
        
        # Combinar y filtrar Nones
        gl_combinados = gl_inicio_raw + gl_fin_raw
        elemento['grados_libertad_global'] = [gl for gl in gl_combinados if gl is not None]

    # Filtrar solo los GL libres (no nulos) para la lista oficial
    st.session_state.grados_libertad_info = [info for info in all_gl_info if info['numero'] is not None]
    
# --- Funciones de Resolución de Sistemas ---

def resolver_sistema_dinamico():
    """Resolver el problema de autovalores para análisis dinámico"""
    if not st.session_state.elementos or not st.session_state.grados_libertad_info:
        return None
    
    try:
        max_gl = len(st.session_state.grados_libertad_info)
        K_global = np.zeros((max_gl, max_gl))
        M_global = np.zeros((max_gl, max_gl))
        
        # Ensamblar matrices K y M globales
        for elemento in st.session_state.elementos:
            if elemento['id'] in st.session_state.matrices_elementos:
                matriz_k_num = np.array(st.session_state.matrices_elementos[elemento['id']].get('numerica', []))
                matriz_m_num = np.array(st.session_state.matrices_elementos[elemento['id']].get('masa_global', []))
                
                gl = elemento['grados_libertad_global']
                
                if matriz_k_num.shape[0] != len(gl) or matriz_m_num.shape[0] != len(gl):
                    st.error(f"Error en Elemento {elemento['id']}: Inconsistencia de GL. Matriz K/M tiene {matriz_k_num.shape[0]} GLs pero el elemento tiene {len(gl)} GLs libres.")
                    continue

                for i, gl_i in enumerate(gl):
                    for j, gl_j in enumerate(gl):
                        K_global[gl_i-1, gl_j-1] += matriz_k_num[i, j]
                        M_global[gl_i-1, gl_j-1] += matriz_m_num[i, j]
        
        # Identificar DOF restringidos (desde el paso 9 dinámico)
        dof_restringidos_idx = [] # base-0
        for gl_num, restringido in st.session_state.condiciones_contorno_dinamica.items():
            if restringido:
                dof_restringidos_idx.append(gl_num - 1)
        
        dof_libres_idx = [i for i in range(max_gl) if i not in dof_restringidos_idx]
        
        if not dof_libres_idx:
            st.error("No hay grados de libertad libres. No se puede calcular el sistema dinámico.")
            return None

        # Extraer submatrices para DOF libres
        K_libre = K_global[np.ix_(dof_libres_idx, dof_libres_idx)]
        M_libre = M_global[np.ix_(dof_libres_idx, dof_libres_idx)]
        
        # Resolver problema de autovalores: K * Φ = λ * M * Φ
        eigenvalues, eigenvectors = eig(K_libre, M_libre)
        
        # Ordenar por frecuencia ascendente (usar .real para autovalores)
        idx = np.argsort(eigenvalues.real)
        eigenvalues_sorted = eigenvalues.real[idx]
        eigenvectors_sorted = eigenvectors.real[:, idx]
        
        # Filtrar valores negativos o muy pequeños (numéricamente inestables)
        min_eigenvalue_threshold = 1e-9
        valid_indices = eigenvalues_sorted > min_eigenvalue_threshold
        
        eigenvalues_valid = eigenvalues_sorted[valid_indices]
        eigenvectors_valid = eigenvectors_sorted[:, valid_indices]

        # Calcular frecuencias naturales (ω² → ω)
        frecuencias_rad = np.sqrt(eigenvalues_valid)
        frecuencias_hz = frecuencias_rad / (2 * np.pi)
        
        # Modos normalizados (por ejemplo, normalización de masa)
        # O simplemente normalizar a la amplitud máxima = 1
        modos_normalizados = eigenvectors_valid / np.max(np.abs(eigenvectors_valid), axis=0)
        
        # Mapear los índices libres (0, 1, 2...) a los números de GL (1, 2, 3...)
        dof_libres_globales = [dof_libres_idx[i] + 1 for i in range(len(dof_libres_idx))]

        return {
            'exito': True,
            'frecuencias_rad': frecuencias_rad,
            'frecuencias_hz': frecuencias_hz,
            'eigenvalues': eigenvalues_valid,
            'eigenvectors': modos_normalizados,
            'dof_libres': dof_libres_globales, # GLs base-1
            'dof_restringidos': [i + 1 for i in dof_restringidos_idx], # GLs base-1
            'K_global': K_global,
            'M_global': M_global,
            'K_libre': K_libre,
            'M_libre': M_libre
        }
    except Exception as e:
        st.error(f"Error en el cálculo dinámico: {str(e)}")
        return None

def resolver_sistema():
    """Resolver el sistema de ecuaciones para análisis estático"""
    if not st.session_state.elementos or not st.session_state.grados_libertad_info:
        return None
    
    try:
        max_gl = len(st.session_state.grados_libertad_info)
        K_global = np.zeros((max_gl, max_gl))
        
        for elemento in st.session_state.elementos:
            if elemento['id'] in st.session_state.matrices_elementos:
                matriz_num = np.array(st.session_state.matrices_elementos[elemento['id']]['numerica'])
                gl = elemento['grados_libertad_global']
                
                if matriz_num.shape[0] != len(gl):
                    st.error(f"Error en Elemento {elemento['id']}: Inconsistencia de GL. Matriz K tiene {matriz_num.shape[0]} GLs pero el elemento tiene {len(gl)} GLs libres.")
                    continue

                for i, gl_i in enumerate(gl):
                    for j, gl_j in enumerate(gl):
                        K_global[gl_i-1, gl_j-1] += matriz_num[i, j]
        
        F = np.zeros(max_gl)
        U = np.zeros(max_gl)
        
        incognitas_u_idx = [] # base-0
        conocidos_u_idx = [] # base-0
        
        for i, info in enumerate(st.session_state.grados_libertad_info):
            gl_idx = info['numero'] - 1 # base-0
            if info['fuerza_conocida']:
                F[gl_idx] = info['valor_fuerza']
            if info['desplazamiento_conocido']:
                U[gl_idx] = info['valor_desplazamiento']
                conocidos_u_idx.append(gl_idx)
            else:
                incognitas_u_idx.append(gl_idx)

        if incognitas_u_idx:
            K_uu = K_global[np.ix_(incognitas_u_idx, incognitas_u_idx)]
            K_uk = K_global[np.ix_(incognitas_u_idx, conocidos_u_idx)] if conocidos_u_idx else np.zeros((len(incognitas_u_idx), 0))
            
            F_u = F[incognitas_u_idx]
            U_k = U[conocidos_u_idx] if conocidos_u_idx else np.array([])
            
            F_efectivo = F_u - (K_uk @ U_k if conocidos_u_idx else 0)
            
            try:
                U_u = np.linalg.solve(K_uu, F_efectivo)
                
                for i, idx in enumerate(incognitas_u_idx):
                    U[idx] = U_u[i]
            except np.linalg.LinAlgError:
                st.error("Error: La matriz de rigidez es singular. El sistema es inestable o tiene movimientos de cuerpo rígido. Verifique sus condiciones de contorno.")
                return None
        
        F_calculado = K_global @ U
        
        # Restaurar fuerzas conocidas (reacciones)
        for i, info in enumerate(st.session_state.grados_libertad_info):
             gl_idx = info['numero'] - 1
             if not info['desplazamiento_conocido']: # Si el desplazamiento era incógnita
                 F_calculado[gl_idx] = F[gl_idx] # La fuerza era conocida (aplicada)

        
        return {
            'K_global': K_global,
            'desplazamientos': U,
            'fuerzas': F_calculado,
            'determinante': np.linalg.det(K_global), # Nota: K_global puede ser singular si hay BCs
            'exito': True
        }
        
    except Exception as e:
        st.error(f"Error resolviendo el sistema estático: {e}")
        return None

# --- Funciones de Modo Interactivo (de V4.7) ---

def crear_grafico_interactivo_moderno():
    """Crear un gráfico interactivo con estilo moderno"""
    fig = go.Figure()

    # Configurar aspecto moderno del gráfico
    fig.update_layout(
        title=dict(
            text="Editor Interactivo de Estructura",
            font=dict(family="Inter, sans-serif", size=20, color="#111827", weight=700),
            x=0.5,
            xanchor='center'
        ),
        xaxis=dict(
            title=dict(text="X [m]", font=dict(family="Inter, sans-serif", size=14, color="#374151", weight=600)),
            showgrid=True,
            gridcolor="#E5E7EB",
            gridwidth=1,
            zeroline=True,
            zerolinecolor="#9CA3AF",
            zerolinewidth=2,
            range=[-10, 10],
            tickfont=dict(family="Inter, sans-serif", size=12, color="#6B7280"),
            linecolor="#E5E7EB",
            mirror=True
        ),
        yaxis=dict(
            title=dict(text="Y [m]", font=dict(family="Inter, sans-serif", size=14, color="#374151", weight=600)),
            showgrid=True,
            gridcolor="#E5E7EB",
            gridwidth=1,
            zeroline=True,
            zerolinecolor="#9CA3AF",
            zerolinewidth=2,
            range=[-10, 10],
            scaleanchor="x",
            scaleratio=1,
            tickfont=dict(family="Inter, sans-serif", size=12, color="#6B7280"),
            linecolor="#E5E7EB",
            mirror=True
        ),
        showlegend=True,
        legend=dict(
            font=dict(family="Inter, sans-serif", size=12, color="#374151", weight=500),
            bgcolor='rgba(255,255,255,0.9)',
            bordercolor='#E5E7EB',
            borderwidth=1,
            x=1,
            y=1,
            xanchor='right',
            yanchor='top'
        ),
        height=600,
        margin=dict(l=60, r=60, t=80, b=60),
        plot_bgcolor='white',
        paper_bgcolor='white',
        font=dict(family="Inter, sans-serif")
    )

    # Añadir nodos existentes con estilo moderno
    nodos_fijos_x = []
    nodos_fijos_y = []
    nodos_fijos_text = []

    nodos_libres_x = []
    nodos_libres_y = []
    nodos_libres_text = []

    for nodo in st.session_state.nodos_interactivos:
        if nodo['tipo'] == 'fijo':
            nodos_fijos_x.append(nodo['x'])
            nodos_fijos_y.append(nodo['y'])
            nodos_fijos_text.append(f"Nodo {nodo['id']}<br>({nodo['x']:.1f}, {nodo['y']:.1f})<br>Tipo: Fijo")
        else:
            nodos_libres_x.append(nodo['x'])
            nodos_libres_y.append(nodo['y'])
            nodos_libres_text.append(f"Nodo {nodo['id']}<br>({nodo['x']:.1f}, {nodo['y']:.1f})<br>Tipo: Libre")

    # Añadir nodos fijos con estilo moderno
    if nodos_fijos_x:
        fig.add_trace(go.Scatter(
            x=nodos_fijos_x,
            y=nodos_fijos_y,
            mode='markers+text',
            marker=dict(
                size=12,
                color='#DC2626',
                line=dict(width=2, color='#991B1B'),
                symbol='circle'
            ),
            text=[f"{nodo['id']}" for nodo in st.session_state.nodos_interactivos if nodo['tipo'] == 'fijo'],
            textposition="middle center",
            textfont=dict(size=10, color='white', family="Inter, sans-serif", weight=700),
            hoverinfo='text',
            hovertext=nodos_fijos_text,
            name='Nodos Fijos',
            hovertemplate='<b>%{hovertext}</b><extra></extra>'
        ))

    # Añadir nodos libres con estilo moderno
    if nodos_libres_x:
        fig.add_trace(go.Scatter(
            x=nodos_libres_x,
            y=nodos_libres_y,
            mode='markers+text',
            marker=dict(
                size=12,
                color='#2563EB',
                line=dict(width=2, color='#1D4ED8'),
                symbol='circle'
            ),
            text=[f"{nodo['id']}" for nodo in st.session_state.nodos_interactivos if nodo['tipo'] == 'libre'],
            textposition="middle center",
            textfont=dict(size=10, color='white', family="Inter, sans-serif", weight=700),
            hoverinfo='text',
            hovertext=nodos_libres_text,
            name='Nodos Libres',
            hovertemplate='<b>%{hovertext}</b><extra></extra>'
        ))

    # Añadir elementos con estilo moderno
    for elemento in st.session_state.elementos_interactivos:
        nodo_inicio = next((n for n in st.session_state.nodos_interactivos if n['id'] == elemento['nodo_inicio']), None)
        nodo_fin = next((n for n in st.session_state.nodos_interactivos if n['id'] == elemento['nodo_fin']), None)

        if nodo_inicio and nodo_fin:
            # Calcular punto medio para etiqueta
            mid_x = (nodo_inicio['x'] + nodo_fin['x']) / 2
            mid_y = (nodo_inicio['y'] + nodo_fin['y']) / 2
            longitud = calcular_longitud_elemento(nodo_inicio, nodo_fin)

            # Añadir etiqueta de elemento con estilo moderno
            fig.add_trace(go.Scatter(
                x=[mid_x],
                y=[mid_y],
                mode='text',
                text=[f"E{elemento['id']}"],
                textposition="middle center",
                textfont=dict(size=10, color='#111827', family="Inter, sans-serif", weight=600),
                hoverinfo='skip',
                showlegend=False
            ))

            # Añadir elemento con estilo moderno
            fig.add_trace(go.Scatter(
                x=[nodo_inicio['x'], nodo_fin['x']],
                y=[nodo_inicio['y'], nodo_fin['y']],
                mode='lines',
                line=dict(width=4, color='#000000'),
                name=f"{st.session_state.tipo_elemento.title()} {elemento['id']}",
                hoverinfo='text',
                hovertext=f"<b>{st.session_state.tipo_elemento.title()} {elemento['id']}</b><br>Nodo {nodo_inicio['id']} → Nodo {nodo_fin['id']}<br>Longitud: {longitud:.3f} m",
                showlegend=True,
                hovertemplate='%{hovertext}<extra></extra>'
            ))

    # Configurar interactividad
    fig.update_layout(
        dragmode='pan',
        clickmode='event+select',
        hovermode='closest'
    )

    return fig

def agregar_nodo_interactivo(x, y, tipo='libre'):
    nodo_id = len(st.session_state.nodos_interactivos) + 1
    # Usar la función simple de GL para el modo interactivo
    gl_globales = calcular_grados_libertad_globales(nodo_id)
    
    nuevo_nodo = {
        'id': nodo_id,
        'x': x,
        'y': y,
        'tipo': tipo,
        'grados_libertad_globales': gl_globales
    }
    
    st.session_state.nodos_interactivos.append(nuevo_nodo)
    return nodo_id

def agregar_elemento_interactivo(nodo_inicio_id, nodo_fin_id):
    if nodo_inicio_id == nodo_fin_id:
        return None
    
    for elem in st.session_state.elementos_interactivos:
        if (elem['nodo_inicio'] == nodo_inicio_id and elem['nodo_fin'] == nodo_fin_id) or \
            (elem['nodo_inicio'] == nodo_fin_id and elem['nodo_fin'] == nodo_inicio_id):
            return None
    
    elemento_id = len(st.session_state.elementos_interactivos) + 1
    
    nodo_inicio = next((n for n in st.session_state.nodos_interactivos if n['id'] == nodo_inicio_id), None)
    nodo_fin = next((n for n in st.session_state.nodos_interactivos if n['id'] == nodo_fin_id), None)
    
    if not nodo_inicio or not nodo_fin:
        return None
    
    gl_globales = nodo_inicio['grados_libertad_globales'] + nodo_fin['grados_libertad_globales']
    
    nuevo_elemento = {
        'id': elemento_id,
        'nodo_inicio': nodo_inicio_id,
        'nodo_fin': nodo_fin_id,
        'grados_libertad_global': gl_globales,
        'tipo': st.session_state.tipo_elemento.title(),
        'material': None,
        'tipo_seccion': None,
        'parametros_seccion': {}
    }
    
    st.session_state.elementos_interactivos.append(nuevo_elemento)
    return elemento_id

def eliminar_nodo_interactivo(nodo_id):
    st.session_state.nodos_interactivos = [n for n in st.session_state.nodos_interactivos if n['id'] != nodo_id]
    # Re-indexar nodos
    for i, nodo in enumerate(st.session_state.nodos_interactivos):
        nodo['id'] = i + 1
        nodo['grados_libertad_globales'] = calcular_grados_libertad_globales(i + 1)
    # Eliminar elementos conectados y re-indexar
    st.session_state.elementos_interactivos = [e for e in st.session_state.elementos_interactivos 
                                        if e['nodo_inicio'] != nodo_id and e['nodo_fin'] != nodo_id]
    st.rerun()

def eliminar_elemento_interactivo(elemento_id):
    st.session_state.elementos_interactivos = [e for e in st.session_state.elementos_interactivos if e['id'] != elemento_id]
    # Re-indexar elementos
    for i, elemento in enumerate(st.session_state.elementos_interactivos):
        elemento['id'] = i + 1
    st.rerun()

def transferir_datos_interactivos():
    """Transfiere datos del modo interactivo al modo manual y avanza al paso de definición de elementos."""
    st.session_state.nodos = st.session_state.nodos_interactivos.copy()
    st.session_state.num_nodos = len(st.session_state.nodos)
    st.session_state.num_fijos = sum(1 for n in st.session_state.nodos if n['tipo'] == 'fijo')
    st.session_state.num_libres = st.session_state.num_nodos - st.session_state.num_fijos
    
    st.session_state.elementos = []
    st.session_state.matrices_elementos = {}
    st.session_state.num_elementos = len(st.session_state.elementos_interactivos)
    
    # Pre-poblar st.session_state.elementos desde interactivos
    for elem_interactivo in st.session_state.elementos_interactivos:
        st.session_state.elementos.append(elem_interactivo.copy())

    st.session_state.step = 8  # Ir directamente a definición de elementos (Paso 8)
    st.rerun()

# --- Funciones de Formateo de Tablas y UI ---

def crear_tabla_nodos():
    """Crear tabla de nodos con coordenadas y grados de libertad"""
    if not st.session_state.nodos:
        return pd.DataFrame()

    nodos_data = []
    for nodo in st.session_state.nodos:
        gl_str = ", ".join([f"GL{gl}" for gl in nodo['grados_libertad_globales']])
        if not gl_str:
            gl_str = "Fijo (0)"
            
        nodos_data.append({
            'ID': nodo['id'],
            'Tipo': nodo.get('tipo', 'Libre').title(),
            'X [m]': f"{nodo['x']:.3f}",
            'Y [m]': f"{nodo['y']:.3f}",
            'Grados de Libertad': gl_str,
        })

    return pd.DataFrame(nodos_data)

def crear_tabla_conectividad():
    """Crear tabla de conectividad de elementos con densidad para análisis dinámico"""
    if not st.session_state.elementos:
        return pd.DataFrame()

    conectividad_data = []
    for elem in st.session_state.elementos:
        
        # Manejar secciones (de V4.7)
        tipo_seccion_val = elem.get('tipo_seccion')
        if tipo_seccion_val:
            seccion_str = tipo_seccion_val.replace('_', ' ').title()
        else:
            seccion_str = "No Definida"
        
        # Asegurar que 'inercia' exista
        if 'inercia' not in elem:
            elem['inercia'] = calcular_momento_inercia(elem.get('tipo_seccion'), elem.get('parametros_seccion', {}))
        
        densidad_str = f"{elem.get('densidad', 0.0):.2f}"
        
        data_elem = {
            'Elemento': elem.get('id', 'N/A'),
            'Nodo Inicio': elem.get('nodo_inicio', 'N/A'),
            'Nodo Fin': elem.get('nodo_fin', 'N/A'),
            'Material': elem.get('material', 'No Definido'),
            'Sección': seccion_str,
            'Área [m²]': f"{elem.get('area', 0.0):.6f}",
        }
        
        if st.session_state.tipo_elemento in ["viga", "viga_portico"]:
             data_elem['Inercia [m⁴]'] = f"{elem.get('inercia', 0.0):.6e}"
        
        if st.session_state.tipo_analisis == "dinamico":
             data_elem['Densidad [kg/m³]'] = densidad_str
        
        data_elem.update({
            'Longitud [m]': f"{elem.get('longitud', 0.0):.3f}",
            'Ángulo β [°]': f"{math.degrees(elem.get('beta', 0.0)):.4f}",
            'GL Globales': str(elem.get('grados_libertad_global', '[]'))
        })
        
        conectividad_data.append(data_elem)

    return pd.DataFrame(conectividad_data)

def crear_tabla_modos_completa():
    """Crear tabla completa con todos los modos y sus amplitudes en todos los DOF,
       formateada como la imagen de Excel."""
    if not st.session_state.resultados_dinamicos:
        return pd.DataFrame()
    
    resultado_din = st.session_state.resultados_dinamicos
    num_modos = len(resultado_din['frecuencias_hz'])
    dof_libres_nums = resultado_din['dof_libres'] # Estos son los números de GL (índices base-1)
    
    # Crear estructura de datos
    datos_modos = []
    
    # Ordenar info GL por número de GL para asegurar el orden correcto
    gl_info_map = {info['numero']: info for info in st.session_state.grados_libertad_info}
    
    # Crear lista de info de GL ordenada según dof_libres_nums
    gl_info_ordenada = []
    for gl_num in dof_libres_nums:
        if gl_num in gl_info_map:
            gl_info_ordenada.append(gl_info_map[gl_num])
        
    
    # --- Encabezados ---
    row_modo = {'Parámetro': 'Modo'}
    row_omega_sq = {'Parámetro': 'ω²'}
    row_hz = {'Parámetro': 'Frecuencia Natural [Hz]'}
    
    for i in range(num_modos):
        modo_str = f'Modo {i+1}'
        row_modo[modo_str] = i + 1
        row_omega_sq[modo_str] = f"{resultado_din['eigenvalues'][i]:.3e}"
        row_hz[modo_str] = f"{resultado_din['frecuencias_hz'][i]:.1f}"
        
    datos_modos.append(row_modo)
    datos_modos.append(row_omega_sq)
    datos_modos.append(row_hz)
    
    # --- Filas para cada DOF libre ---
    
    for i, info_gl in enumerate(gl_info_ordenada):
        gl_num = info_gl['numero']
        # El PDF usa 'a4', 'a5', etc. Lo replicamos
        row_dof = {'Parámetro': f"a{gl_num}"}
        
        for modo_idx in range(num_modos):
            # i es el índice del eigenvector (0 a N_libres-1)
            amplitud = resultado_din['eigenvectors'][i, modo_idx]
            row_dof[f'Modo {modo_idx+1}'] = f"{amplitud:.3e}"
        
        datos_modos.append(row_dof)
    
    return pd.DataFrame(datos_modos)

# --- Funciones de Visualización (Estática y Dinámica) ---

def mostrar_barra_progreso():
    """Mostrar barra de progreso estilo web moderno (CORREGIDA)"""
    if st.session_state.step == 0:
        return
    
    # Definir la secuencia de pasos base
    pasos_base = [
        (0, "Análisis"),
        (1, "Elemento"),
        (2, "Usuario"),
        (3, "Modo Entrada")
    ]
    
    # Pasos específicos del modo
    if st.session_state.modo == "manual":
        pasos_modo = [
            (4, "Nº Nodos"),
            (5, "Coordenadas"),
            (6, "Clasificación Nodos"),
            (7, "Nº Elementos")
        ]
    else: # Modo Interactivo (salta de 3 a 8)
        pasos_modo = [] 

    # Pasos finales (comunes a ambos modos)
    pasos_finales_comunes = [
        (8, "Def. Elementos")
    ]
    
    # Pasos de análisis (Estático vs Dinámico)
    if st.session_state.tipo_analisis == "dinamico":
        pasos_analisis = [
            (9, "Cond. Contorno"),
            (10, "Resultados") # El Paso 10 ahora es "Resultados"
        ]
    else: # Estático
        pasos_analisis = [
            (9, "Incógnitas"),
            (10, "Resultados") # El Paso 10 ahora es "Resultados"
        ]
        
    # Combinar todas las listas de pasos
    pasos_tuplas = pasos_base + pasos_modo + pasos_finales_comunes + pasos_analisis
    
    # Mapear el número de paso (st.session_state.step) a su índice en la lista de pasos
    pasos_map = {step_num: i for i, (step_num, name) in enumerate(pasos_tuplas)}
    
    # Obtener el índice actual (ej: si step=9, el índice_actual es 8)
    paso_actual_idx = -1
    if st.session_state.step in pasos_map:
        paso_actual_idx = pasos_map[st.session_state.step]
    
    pasos_nombres = [name for (step_num, name) in pasos_tuplas]
    
    st.markdown("""
    <div class='progress-bar'>
        <div style='max-width: 1200px; margin: 0 auto; padding: 0 2rem;'>
            <div style='display: flex; justify-content: space-between; align-items: center; margin-bottom: 1rem;'>
                <h1 style='margin: 0; font-size: 1.8rem;'>Análisis Estructural - Método de Matrices</h1>
                <div style='display: flex; gap: 1rem;'>
    """, unsafe_allow_html=True)
    
    # Botones de control en la barra
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("🔄 Reiniciar", key="reset_top"):
            reset_app()
    with col2:
        # El botón "Anterior" no debe aparecer en el paso 1 o 0
        if st.session_state.step > 1:
            if st.button("← Anterior", key="prev_top"):
                prev_step()
    
    st.markdown("</div></div>", unsafe_allow_html=True)
    
    # Mostrar pasos con círculos
    progress_html = "<div class='progress-steps'>"
    
    for i, paso_nombre in enumerate(pasos_nombres):
        # Usamos el índice (i) para la comparación
        if i < paso_actual_idx:
            circle_class = "completed"
            circle_text = "✓"
        elif i == paso_actual_idx:
            circle_class = "current"
            circle_text = str(i + 1)
        else:
            circle_class = "pending"
            circle_text = str(i + 1)
        
        progress_html += f"""
        <div class='progress-step'>
            <div class='step-circle {circle_class}'>
                {circle_text}
            </div>
        """
        
        if i < len(pasos_nombres) - 1:
            line_class = "completed" if i < paso_actual_idx else ""
            progress_html += f"<div class='step-line {line_class}'></div>"
        
        progress_html += "</div>"
    
    progress_html += "</div>"
    
    # Mostrar paso actual
    if paso_actual_idx != -1 and paso_actual_idx < len(pasos_nombres):
        progress_html += f"""
            <div style='text-align: center; margin-top: 1rem;'>
                <div style='font-size: 1.1rem; font-weight: 600; color: var(--gray-800);'>
                    Paso {paso_actual_idx + 1}: {pasos_nombres[paso_actual_idx]}
                </div>
            </div>
        </div>
        """
    
    st.markdown(progress_html, unsafe_allow_html=True)

def mostrar_sidebar_mejorado():
    """Mostrar sidebar mejorado solo cuando hay tipo seleccionado (de V4.7)"""
    if st.session_state.step == 0:
        return
    
    st.markdown('<div class="show-sidebar">', unsafe_allow_html=True)
    
    with st.sidebar:
        st.markdown("### Progreso del Análisis")
        
        # Definir la secuencia de pasos base
        pasos_base = [
            (0, "Análisis"),
            (1, "Elemento"),
            (2, "Usuario"),
            (3, "Modo Entrada")
        ]
        
        # Pasos específicos del modo
        if st.session_state.modo == "manual":
            pasos_modo = [
                (4, "Nº Nodos"),
                (5, "Coordenadas"),
                (6, "Clasificación Nodos"),
                (7, "Nº Elementos")
            ]
        else: # Modo Interactivo (salta de 3 a 8)
            pasos_modo = [] 

        # Pasos finales (comunes a ambos modos)
        pasos_finales_comunes = [
            (8, "Def. Elementos")
        ]
        
        # Pasos de análisis (Estático vs Dinámico)
        if st.session_state.tipo_analisis == "dinamico":
            pasos_analisis = [
                (9, "Cond. Contorno"),
                (10, "Resultados") # El Paso 10 ahora es "Resultados"
            ]
        else: # Estático
            pasos_analisis = [
                (9, "Incógnitas"),
                (10, "Resultados") # El Paso 10 ahora es "Resultados"
            ]
            
        pasos_tuplas = pasos_base + pasos_modo + pasos_finales_comunes + pasos_analisis
        
        for step_num, paso_nombre in pasos_tuplas:
            if step_num == st.session_state.step:
                st.markdown(f"**→ {paso_nombre}**")
            elif step_num < st.session_state.step:
                st.markdown(f"✓ {paso_nombre}")
            else:
                st.markdown(f"⏳ {paso_nombre}")
        
        st.divider()
        
        st.markdown("### Información del Proyecto")
        
        if st.session_state.usuario_nombre:
            st.markdown(f"**Usuario:** {st.session_state.usuario_nombre}")
        
        if st.session_state.tipo_analisis:
            st.markdown(f"**Análisis:** {st.session_state.tipo_analisis.title()}")
            
        if st.session_state.tipo_elemento:
            st.markdown(f"**Elemento:** {st.session_state.tipo_elemento.replace('_', ' ').title()}")
        
        if st.session_state.modo:
            st.markdown(f"**Modo:** {st.session_state.modo.capitalize()}")
        
        if st.session_state.nodos:
            st.markdown(f"**Nodos:** {len(st.session_state.nodos)}")
        
        if st.session_state.elementos:
            st.markdown(f"**Elementos:** {len(st.session_state.elementos)}")
        
        if st.session_state.grados_libertad_info:
            st.markdown(f"**Grados de Libertad:** {len(st.session_state.grados_libertad_info)}")
        
        st.markdown(f"**Fecha:** {datetime.now().strftime('%d/%m/%Y')}")

def mostrar_matriz_formateada_moderna(matriz, titulo="Matriz", es_simbolica=True):
    """Mostrar matriz en formato tabla con estilo moderno (de V4.7)"""
    if matriz is None or len(matriz) == 0:
        st.warning(f"⚠️ Matriz '{titulo}' vacía o no disponible.")
        return

    st.markdown(f"##### {titulo}")

    if es_simbolica:
        df = pd.DataFrame(matriz)
    else:
        # Formatear números para legibilidad
        matriz_formateada = []
        for fila in matriz:
            fila_formateada = [f"{valor:.3e}" for valor in fila]
            matriz_formateada.append(fila_formateada)
        
        df = pd.DataFrame(matriz_formateada)
        if len(matriz) > 0:
            df.index = [f"GL {i+1}" for i in range(len(matriz))]
            df.columns = [f"GL {j+1}" for j in range(len(matriz[0]))]

    st.dataframe(df, use_container_width=True)

def hermite_interpolation(v1, theta1, v2, theta2, L, num_points=50):
    """
    Calcula la forma de la deflexión de una viga usando polinomios de Hermite.
    Retorna las coordenadas locales (x, v) de la curva.
    """
    x_local = np.linspace(0, L, num_points)
    s = x_local / L  # Coordenada normalizada

    # Funciones de forma cúbicas de Hermite
    H1 = 2*s**3 - 3*s**2 + 1
    H2 = (s**3 - 2*s**2 + s) * L
    H3 = -2*s**3 + 3*s**2
    H4 = (s**3 - s**2) * L

    # Deflexión transversal v(x) en coordenadas locales
    v_local = H1*v1 + H2*theta1 + H3*v2 + H4*theta2
    
    return x_local, v_local

def visualizar_estructura_moderna(mostrar_deformada=False, factor_escala=10):
    """Visualizar la estructura estática con estilo moderno (de V4.7)"""
    if not st.session_state.nodos:
        st.warning("No hay nodos para visualizar")
        return None
    
    plt.style.use('default')
    fig, ax = plt.subplots(figsize=(8, 8), facecolor='white')
    ax.set_facecolor('white')

    nodos_deformados = []
    all_x = [nodo['x'] for nodo in st.session_state.nodos]
    all_y = [nodo['y'] for nodo in st.session_state.nodos]

    if mostrar_deformada and st.session_state.resultados:
        for nodo in st.session_state.nodos:
            nodo_deformado = nodo.copy()
            dx, dy, dtheta = 0, 0, 0
            
            # Encontrar los GL de este nodo
            gls_nodo = nodo['grados_libertad_globales']
            
            if nodo['tipo'] != 'fijo' and gls_nodo:
                desplazamientos_res = st.session_state.resultados['desplazamientos']
                
                if st.session_state.tipo_elemento == "barra":
                    dx = desplazamientos_res[gls_nodo[0]-1]
                    dy = desplazamientos_res[gls_nodo[1]-1]
                elif st.session_state.tipo_elemento == "viga":
                    dy = desplazamientos_res[gls_nodo[0]-1]
                    dtheta = desplazamientos_res[gls_nodo[1]-1]
                elif st.session_state.tipo_elemento == "viga_portico":
                    dx = desplazamientos_res[gls_nodo[0]-1]
                    dy = desplazamientos_res[gls_nodo[1]-1]
                    dtheta = desplazamientos_res[gls_nodo[2]-1]

            nodo_deformado['x'] += dx * factor_escala
            nodo_deformado['y'] += dy * factor_escala
            nodo_deformado['theta'] = dtheta # Guardar rotación
            nodos_deformados.append(nodo_deformado)
            
            all_x.append(nodo_deformado['x'])
            all_y.append(nodo_deformado['y'])

    # Dibujar original (si se muestra deformada)
    if mostrar_deformada:
        for elemento in st.session_state.elementos:
            nodo_inicio = next((n for n in st.session_state.nodos if n['id'] == elemento['nodo_inicio']), None)
            nodo_fin = next((n for n in st.session_state.nodos if n['id'] == elemento['nodo_fin']), None)
            if nodo_inicio and nodo_fin:
                ax.plot([nodo_inicio['x'], nodo_fin['x']], 
                        [nodo_inicio['y'], nodo_fin['y']], 
                        color='#ced4da', linewidth=2, alpha=0.8, linestyle='--', 
                        label='Estructura Original' if 'Estructura Original' not in ax.get_legend_handles_labels()[1] else "")

    # Dibujar elementos
    for elemento in st.session_state.elementos:
        nodo_inicio = next((n for n in st.session_state.nodos if n['id'] == elemento['nodo_inicio']), None)
        nodo_fin = next((n for n in st.session_state.nodos if n['id'] == elemento['nodo_fin']), None)
        
        if not (nodo_inicio and nodo_fin):
            continue

        if mostrar_deformada and nodos_deformados:
            nodo_inicio_def = next((n for n in nodos_deformados if n['id'] == elemento['nodo_inicio']), None)
            nodo_fin_def = next((n for n in nodos_deformados if n['id'] == elemento['nodo_fin']), None)

            if nodo_inicio_def and nodo_fin_def and st.session_state.tipo_elemento in ["viga", "viga_portico"]:
                U_global_vec = st.session_state.resultados['desplazamientos']
                T = np.eye(6) 
                u_local = np.zeros(6)

                if st.session_state.tipo_elemento == "viga":
                    gl_i = nodo_inicio['grados_libertad_globales']
                    gl_j = nodo_fin['grados_libertad_globales']
                    u_local = np.array([0, U_global_vec[gl_i[0]-1], U_global_vec[gl_i[1]-1], 0, U_global_vec[gl_j[0]-1], U_global_vec[gl_j[1]-1]])
                
                elif st.session_state.tipo_elemento == "viga_portico":
                    T = generar_matriz_transformacion_viga_portico(elemento['beta'])
                    gl_globales_elem = elemento['grados_libertad_global']
                    U_global_elem = np.array([U_global_vec[i-1] for i in gl_globales_elem])
                    u_local = T @ U_global_elem

                v1, theta1 = u_local[1], u_local[2]
                v2, theta2 = u_local[4], u_local[5]

                x_local, v_local_curva = hermite_interpolation(v1, theta1, v2, theta2, elemento['longitud'])
                
                beta = elemento['beta']
                c, s = math.cos(beta), math.sin(beta)
                
                x_global_base = nodo_inicio_def['x'] + x_local * c
                y_global_base = nodo_inicio_def['y'] + x_local * s
                
                x_global_final = x_global_base - v_local_curva * s * factor_escala
                y_global_final = y_global_base + v_local_curva * c * factor_escala

                # --- CORRECCIÓN LÍMITES ---
                all_x.extend(x_global_final)
                all_y.extend(y_global_final)
                # --- FIN CORRECCIÓN ---

                ax.plot(x_global_final, y_global_final,
                        color='#000000', linewidth=3, alpha=0.9,
                        label='Estructura Deformada' if 'Estructura Deformada' not in ax.get_legend_handles_labels()[1] else "")
            
            elif nodo_inicio_def and nodo_fin_def: # Para 'barra'
                ax.plot([nodo_inicio_def['x'], nodo_fin_def['x']], [nodo_inicio_def['y'], nodo_fin_def['y']], 
                        color='#000000', linewidth=3, alpha=0.9,
                        label='Estructura Deformada' if 'Estructura Deformada' not in ax.get_legend_handles_labels()[1] else "")
        else:
            # Dibujar solo original
            ax.plot([nodo_inicio['x'], nodo_fin['x']], [nodo_inicio['y'], nodo_fin['y']], color='#000000', linewidth=3, alpha=0.9)

        # Etiqueta de elemento
        mid_x, mid_y = (nodo_inicio['x'] + nodo_fin['x']) / 2, (nodo_inicio['y'] + nodo_fin['y']) / 2
        ax.text(mid_x, mid_y, f'E{elemento["id"]}', ha='center', va='center', fontsize=9, fontweight='600',
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="black", linewidth=1.5, alpha=0.95), zorder=20)

    # --- CORRECCIÓN LÍMITES: Mover cálculo de límites a DESPUÉS de plotear ---
    x_min, x_max = (min(all_x) if all_x else 0), (max(all_x) if all_x else 1)
    y_min, y_max = (min(all_y) if all_y else 0), (max(all_y) if all_y else 1)
    x_range = x_max - x_min if x_max > x_min else 2
    y_range = y_max - y_min if y_max > y_min else 2
    
    padding_x = 0.1 * x_range
    padding_y = 0.1 * y_range
    
    ax.set_xlim(x_min - padding_x, x_max + padding_x)
    ax.set_ylim(y_min - padding_y, y_max + padding_y)
    # --- FIN CORRECCIÓN LÍMITES ---
    
    # Dibujar nodos
    for i, nodo in enumerate(st.session_state.nodos):
        color_orig = '#DC2626' if nodo['tipo'] == 'fijo' else '#6c757d'
        ax.plot(nodo['x'], nodo['y'], 'o', markersize=12, color=color_orig, zorder=10, markeredgecolor='black')
        
        if mostrar_deformada and nodos_deformados:
            nodo_def = nodos_deformados[i]
            color_def = '#DC2626' if nodo['tipo'] == 'fijo' else '#28a745'
            ax.plot(nodo_def['x'], nodo_def['y'], 'o', markersize=12, color=color_def, zorder=11, markeredgecolor='black')
            ax.plot([nodo['x'], nodo_def['x']], [nodo['y'], nodo_def['y']], color='#6c757d', linestyle=':', linewidth=1.5)
        
        ax.text(nodo['x'], nodo['y'], str(nodo['id']), ha='center', va='center', fontsize=9, fontweight='700', color='white', zorder=12)

    ax.set_xlabel('X [m]', fontsize=12, fontweight='600')
    ax.set_ylabel('Y [m]', fontsize=12, fontweight='600')
    ax.set_title(f'Estructura Deformada (x{factor_escala})' if mostrar_deformada else 'Estructura Original', fontsize=16, fontweight='700', pad=20)
    ax.grid(True, which='both', linestyle='--', linewidth=0.5, color='#adb5bd')
    
    ax.set_aspect('auto')
    
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        unique_labels = dict(zip(labels, handles))
        ax.legend(unique_labels.values(), unique_labels.keys(), loc='upper right', frameon=True, fancybox=True, shadow=True, fontsize=10)
    
    plt.tight_layout()
    return fig

def dibujar_apoyo(ax, x, y, tipo='fijo', color='black', size=10):
    """
    Dibuja un símbolo de apoyo (empotrado/fijo) en las coordenadas dadas.
    """
    if tipo == 'fijo':
        # Base del triángulo
        base_x = [x - size/20, x + size/20, x]
        base_y = [y - size/10, y - size/10, y]
        
        # Triángulo
        triangle = plt.Polygon(list(zip(base_x, base_y)), color=color, fill=False, linewidth=2)
        ax.add_patch(triangle)
        
        # Línea de "tierra"
        line_x = [x - size/15, x + size/15]
        line_y = [y - size/10, y - size/10]
        ax.plot(line_x, line_y, color=color, linewidth=2)
        
        # Rayas de empotramiento
        for i in np.linspace(line_x[0], line_x[1], 4):
            ax.plot([i, i - size/40], [line_y[0], line_y[0] - size/40], color=color, linewidth=1)

def visualizar_modo_dinamico(modo_idx, factor_escala=None, figsize=(8, 8)):
    """
    Visualiza un modo de vibración específico para la estructura,
    con una mejor estética y relación de aspecto. (Versión CORREGIDA)
    
    Args:
        modo_idx (int): Índice del modo a visualizar (0-based).
        factor_escala (float, optional): Factor para escalar la deformada.
                                         Si es None, se escala automáticamente.
        figsize (tuple, optional): Tamaño de la figura (ancho, alto) en pulgadas.
    Returns:
        matplotlib.figure.Figure: La figura de matplotlib con la visualización.
    """
    if not st.session_state.resultados_dinamicos or not st.session_state.resultados_dinamicos.get('exito'):
        st.warning("No hay resultados dinámicos para visualizar.")
        return None

    # 1. Obtener el vector de modo "pequeño" (solo DOFs libres) desde la clave correcta
    eigenvector_matrix_small = np.array(st.session_state.resultados_dinamicos['eigenvectors'])
    
    # Comprobar si modo_idx es válido
    if modo_idx < 0 or modo_idx >= eigenvector_matrix_small.shape[1]:
        st.error(f"Índice de modo inválido: {modo_idx}. Debe estar entre 0 y {eigenvector_matrix_small.shape[1]-1}.")
        return None
        
    modo_shape_small = eigenvector_matrix_small[:, modo_idx]

    # 2. Obtener la lista de DOFs libres (números base-1, ej: [3, 4, 5])
    dof_libres_nums = st.session_state.resultados_dinamicos['dof_libres'] 

    # 3. Crear un mapa de {GL_num: indice_en_vector_pequeño}
    gl_a_idx_map = {gl_num: i for i, gl_num in enumerate(dof_libres_nums)}

    # 4. Crear el vector de desplazamiento "completo" (incluyendo 0s para DOFs fijos)
    num_total_gl = len(st.session_state.grados_libertad_info)
    desplazamientos_modales_actuales_full = np.zeros(num_total_gl)

    # 5. Llenar el vector "completo" con los valores del vector "pequeño"
    for gl_num, i in gl_a_idx_map.items():
        if (gl_num - 1) < num_total_gl: # Verificación de seguridad
            desplazamientos_modales_actuales_full[gl_num - 1] = modo_shape_small[i]
    
    # Obtener el nombre del modo y la frecuencia
    f_modo = st.session_state.resultados_dinamicos['frecuencias_hz'][modo_idx]
    omega_modo = st.session_state.resultados_dinamicos['frecuencias_rad'][modo_idx]
    
    # Crear la figura y los ejes
    fig, ax = plt.subplots(figsize=figsize)
    ax.set_aspect('auto') # Dejar que los límites definan la proporción

    ax.set_xlabel("X [m]")
    ax.set_ylabel("Y [m]")
    ax.grid(True, linestyle='--', alpha=0.6)
    
    # Rango de la estructura original para el ajuste automático del factor de escala
    x_coords_orig = [n['x'] for n in st.session_state.nodos]
    y_coords_orig = [n['y'] for n in st.session_state.nodos]
    
    min_x_orig, max_x_orig = (min(x_coords_orig) if x_coords_orig else 0), (max(x_coords_orig) if x_coords_orig else 0)
    min_y_orig, max_y_orig = (min(y_coords_orig) if y_coords_orig else 0), (max(y_coords_orig) if y_coords_orig else 0)
    
    rango_x_orig = max_x_orig - min_x_orig if max_x_orig != min_x_orig else 1.0
    rango_y_orig = max_y_orig - min_y_orig if max_y_orig != min_y_orig else 1.0
    rango_global_orig = max(rango_x_orig, rango_y_orig, 1.0) # Asegurarse de que no sea 0

    # Para cálculo automático del factor de escala
    max_desplazamiento_nodal = 0.0
    
    gl_info_map = {gl['numero']: gl for gl in st.session_state.grados_libertad_info}
    gl_info_map_by_node = {}
    for gl in st.session_state.grados_libertad_info:
        if gl['nodo'] not in gl_info_map_by_node:
            gl_info_map_by_node[gl['nodo']] = []
        gl_info_map_by_node[gl['nodo']].append(gl)

    for nodo_original in st.session_state.nodos:
        gls_nodo = gl_info_map_by_node.get(nodo_original['id'], [])
        
        for gl_info in gls_nodo:
            gl_num = gl_info['numero']
            despl = desplazamientos_modales_actuales_full[gl_num - 1] 
            
            if gl_num in dof_libres_nums:
                 max_desplazamiento_nodal = max(max_desplazamiento_nodal, abs(despl))

    # Calcular factor de escala automático si no se proporcionó
    if factor_escala is None:
        if max_desplazamiento_nodal > 1e-9: 
            factor_escala = (0.1 * rango_global_orig) / max_desplazamiento_nodal 
            factor_escala = max(1, min(factor_escala, 500)) 
        else:
            factor_escala = 1.0 

    # Dibujar nodos y elementos
    line_styles = {
        'original': {'color': 'blue', 'linestyle': ':', 'linewidth': 1.5, 'label': 'Original'},
        'deformada': {'color': 'darkgreen', 'linestyle': '-', 'linewidth': 2.5, 'label': f'Modo (Escala x{factor_escala:.1f})'}
    }

    # Recopilar coordenadas para los límites del gráfico
    all_x = list(x_coords_orig)
    all_y = list(y_coords_orig)
    
    gl_dir_map_by_node = {}
    for gl in st.session_state.grados_libertad_info:
        if gl['nodo'] not in gl_dir_map_by_node:
            gl_dir_map_by_node[gl['nodo']] = {}
        gl_dir_map_by_node[gl['nodo']][gl['direccion'].lower()] = gl['numero']


    # Dibujar elementos originales y deformados
    for elemento in st.session_state.elementos:
        nodo_inicio_original = next(n for n in st.session_state.nodos if n['id'] == elemento['nodo_inicio'])
        nodo_fin_original = next(n for n in st.session_state.nodos if n['id'] == elemento['nodo_fin'])

        x1_orig, y1_orig = nodo_inicio_original['x'], nodo_inicio_original['y']
        x2_orig, y2_orig = nodo_fin_original['x'], nodo_fin_original['y']

        # Dibujar original
        ax.plot([x1_orig, x2_orig], [y1_orig, y2_orig], **line_styles['original'])

        # Calcular deformada con el factor de escala
        gls_inicio = gl_dir_map_by_node.get(nodo_inicio_original['id'], {})
        gls_fin = gl_dir_map_by_node.get(nodo_fin_original['id'], {})

        dx1 = desplazamientos_modales_actuales_full[gls_inicio.get('x', 1)-1] * factor_escala if 'x' in gls_inicio else 0
        dy1 = desplazamientos_modales_actuales_full[gls_inicio.get('y', 1)-1] * factor_escala if 'y' in gls_inicio else 0
        dtheta1 = desplazamientos_modales_actuales_full[gls_inicio.get('theta', 1)-1] * factor_escala if 'theta' in gls_inicio else 0
        
        dx2 = desplazamientos_modales_actuales_full[gls_fin.get('x', 1)-1] * factor_escala if 'x' in gls_fin else 0
        dy2 = desplazamientos_modales_actuales_full[gls_fin.get('y', 1)-1] * factor_escala if 'y' in gls_fin else 0
        dtheta2 = desplazamientos_modales_actuales_full[gls_fin.get('theta', 1)-1] * factor_escala if 'theta' in gls_fin else 0
        
        x1_def, y1_def = x1_orig + dx1, y1_orig + dy1
        x2_def, y2_def = x2_orig + dx2, y2_orig + dy2

        all_x.extend([x1_def, x2_def])
        all_y.extend([y1_def, y2_def])

        # Interpolación de Hermite para la deformada de vigas/pórticos
        if st.session_state.tipo_elemento in ["viga", "viga_portico"]:
            dx_orig = x2_orig - x1_orig
            dy_orig = y2_orig - y1_orig
            L_orig = np.sqrt(dx_orig**2 + dy_orig**2)
            
            if L_orig < 1e-9: # Evitar división por cero
                ax.plot([x1_def, x2_def], [y1_def, y2_def], **line_styles['deformada'])
                continue

            beta = np.arctan2(dy_orig, dx_orig)
            
            v1_loc_scaled, theta1_loc_scaled, v2_loc_scaled, theta2_loc_scaled = 0, 0, 0, 0
            u1_loc_scaled, u2_loc_scaled = 0, 0
            beta_hermite = 0.0 # Por defecto (para viga)
            
            if st.session_state.tipo_elemento == "viga":
                v1_loc_scaled = dy1
                theta1_loc_scaled = dtheta1
                v2_loc_scaled = dy2
                theta2_loc_scaled = dtheta2
                beta_hermite = 0.0 # Viga asume beta=0
            
            else:  # viga_portico
                despl_global_elem_scaled = np.array([dx1, dy1, dtheta1, dx2, dy2, dtheta2])
                T = generar_matriz_transformacion_viga_portico(beta)
                despl_local_elem_scaled = T @ despl_global_elem_scaled

                u1_loc_scaled = despl_local_elem_scaled[0]
                v1_loc_scaled = despl_local_elem_scaled[1]
                theta1_loc_scaled = despl_local_elem_scaled[2]
                u2_loc_scaled = despl_local_elem_scaled[3]
                v2_loc_scaled = despl_local_elem_scaled[4]
                theta2_loc_scaled = despl_local_elem_scaled[5]
                beta_hermite = beta  # usar el ángulo del elemento

            x_hermite_loc, v_hermite_loc_scaled = hermite_interpolation(v1_loc_scaled, theta1_loc_scaled, v2_loc_scaled, theta2_loc_scaled, L_orig)
            
            x_hermite_glob = x1_orig + x_hermite_loc * np.cos(beta_hermite) - v_hermite_loc_scaled * np.sin(beta_hermite)
            y_hermite_glob = y1_orig + x_hermite_loc * np.sin(beta_hermite) + v_hermite_loc_scaled * np.cos(beta_hermite)
            
            if st.session_state.tipo_elemento == "viga_portico":
                u_hermite_loc_scaled = u1_loc_scaled * (1 - x_hermite_loc/L_orig) + u2_loc_scaled * (x_hermite_loc/L_orig)
                
                x_hermite_glob += u_hermite_loc_scaled * np.cos(beta_hermite)
                y_hermite_glob += u_hermite_loc_scaled * np.sin(beta_hermite)

            # --- CORRECCIÓN LÍMITES ---
            all_x.extend(x_hermite_glob)
            all_y.extend(y_hermite_glob)
            # --- FIN CORRECCIÓN ---

            ax.plot(x_hermite_glob, y_hermite_glob, **line_styles['deformada'])

        else: # Barra: solo dibujar línea recta deformada
            ax.plot([x1_def, x2_def], [y1_def, y2_def], **line_styles['deformada'])

    # --- CORRECCIÓN LÍMITES: Mover cálculo de límites a DESPUÉS de plotear ---
    x_min, x_max = (min(all_x) if all_x else 0), (max(all_x) if all_x else 1)
    y_min, y_max = (min(all_y) if all_y else 0), (max(all_y) if all_y else 1)
    x_range = x_max - x_min if x_max > x_min else 2
    y_range = y_max - y_min if y_max > y_min else 2
    
    padding_x = 0.1 * x_range
    padding_y = 0.1 * y_range
    
    ax.set_xlim(x_min - padding_x, x_max + padding_x)
    ax.set_ylim(y_min - padding_y, y_max + padding_y)
    # --- FIN CORRECCIÓN LÍMITES ---

    # Dibujar nodos originales y deformados (con texto)
    for nodo in st.session_state.nodos:
        idx = nodo['id'] - 1
        x_orig, y_orig = x_coords_orig[idx], y_coords_orig[idx]
        
        gls_nodo = gl_info_map_by_node.get(nodo['id'], [])
        dx, dy = 0, 0
        for gl_info in gls_nodo:
            gl_num = gl_info['numero']
            despl = desplazamientos_modales_actuales_full[gl_num - 1] * factor_escala
            if gl_info['direccion'].lower() == 'x': dx = despl
            elif gl_info['direccion'].lower() == 'y': dy = despl
        
        x_def, y_def = x_orig + dx, y_orig + dy

        # Dibujar Nodo Original
        ax.plot(x_orig, y_orig, 'o', markersize=8, color='blue', markeredgecolor='black', zorder=5)
        ax.text(x_orig, y_orig, f" {nodo['id']}", color='blue', ha='left', va='bottom', fontsize=10, weight='bold')

        # Dibujar Nodo Deformado
        ax.plot(x_def, y_def, 'o', markersize=8, color='lightgreen', markeredgecolor='darkgreen', zorder=5)
        ax.text(x_def, y_def, f" {nodo['id']}'", color='darkgreen', ha='left', va='bottom', fontsize=10, weight='bold')
    
    # Dibujar apoyos (restricciones)
    for nodo in st.session_state.nodos:
        if nodo.get('tipo') == 'fijo':
            dibujar_apoyo(ax, nodo['x'], nodo['y'], 'fijo', 'black', size=rango_global_orig*0.05) # Tamaño relativo

    # Limpiar leyenda de duplicados
    handles, labels = ax.get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    ax.legend(by_label.values(), by_label.keys())
    
    ax.set_title(f"Modo {modo_idx + 1} (f = {f_modo:.2f} Hz, Escala x{factor_escala:.1f})", fontsize=16)
    fig.tight_layout()
    
    return fig

# --- Funciones de Reporte (PDF) ---

def generar_pdf_reporte_dinamico():
    """Generar reporte PDF para análisis dinámico con estilo Excel y todos los gráficos de modos."""
    if not st.session_state.resultados_dinamicos:
        return None
    
    try:
        # Crear documento PDF
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4),
                                rightMargin=inch/2, leftMargin=inch/2,
                                topMargin=inch/2, bottomMargin=inch/2)
        story = []
        styles = getSampleStyleSheet()
        
        # --- Estilos (Todos Azules) ---
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#2d3748'), spaceAfter=30, alignment=1)
        section_style = ParagraphStyle('CustomSection', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#4472C4'), spaceAfter=12, spaceBefore=12)
        
        # Esquema de color Azul (para TODAS las tablas)
        style_header_color = colors.HexColor('#4472C4')
        style_data_color = colors.HexColor('#D9E1F2')
        style_header_font_color = colors.whitesmoke
        style_data_font_color = colors.black

        story.append(Paragraph("ANÁLISIS DINÁMICO - ESTRUCTURAS", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        resultado = st.session_state.resultados_dinamicos
        
        # --- 1. TABLA COMPLETA DE MODOS (Principal) ---
        story.append(Paragraph("1. FORMAS MODALES (EIGENVECTORES)", section_style))
        df_modos = crear_tabla_modos_completa()
        data_modos_full = [list(df_modos.columns)] + df_modos.values.tolist()
        
        col_widths = [1.5*inch] + [0.6*inch]*(len(df_modos.columns)-1)
        if sum(col_widths) > (doc.width):
             col_widths = [1.0*inch] + [ (doc.width - 1.0*inch) / (len(df_modos.columns)-1) ] * (len(df_modos.columns)-1)

        table_modos_full = Table(data_modos_full, colWidths=col_widths)
        # --- CORRECCIÓN DE COLOR ---
        table_modos_full.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), style_header_color), # Encabezado (Modo)
            ('BACKGROUND', (0, 1), (-1, 2), style_header_color), # Encabezado (w2, f)
            ('BACKGROUND', (0, 3), (-1, -1), style_data_color), # Datos
            ('BACKGROUND', (0, 0), (0, -1), style_data_color), # Columna Parámetro
            ('TEXTCOLOR', (0, 0), (-1, 2), style_header_font_color), # Texto encabezado
            ('TEXTCOLOR', (0, 3), (-1, -1), style_data_font_color), # Texto datos
            ('TEXTCOLOR', (0, 0), (0, -1), style_data_font_color), # Texto params
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table_modos_full)
        story.append(PageBreak())

        # --- 2. GRÁFICOS DE TODOS LOS MODOS (Uno por página) ---
        story.append(Paragraph("2. VISUALIZACIÓN DE MODOS", section_style))
        num_modos = len(resultado['frecuencias_hz'])
        
        factor_escala_reporte = None # Usar automático
        
        for i in range(num_modos):
            # --- CORRECCIÓN DE ORDEN: PageBreak ANTES del título y gráfico ---
            story.append(PageBreak()) 
            
            fig_modo = visualizar_modo_dinamico(i, factor_escala_reporte, figsize=(8,8)) # Forzar 8x8
            
            if fig_modo:
                img_buffer = io.BytesIO()
                fig_modo.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
                img_buffer.seek(0)
                
                img_pdf = Image(img_buffer)
                img_pdf.drawHeight = 7 * inch 
                img_pdf.drawWidth = 7 * inch
                img_pdf.hAlign = 'CENTER'
                
                # Título y gráfico JUNTOS
                story.append(Paragraph(f"Modo {i+1} (f = {resultado['frecuencias_hz'][i]:.2f} Hz)", styles['Heading3']))
                story.append(img_pdf)
                
                plt.close(fig_modo)

        story.append(PageBreak())

        # --- 3. TABLAS DE NODOS Y CONECTIVIDAD (Una por página) ---
        story.append(Paragraph("3. INFORMACIÓN DE NODOS Y ELEMENTOS", section_style))
        
        df_nodos = crear_tabla_nodos()
        if not df_nodos.empty:
            data_nodos = [list(df_nodos.columns)] + df_nodos.values.tolist()
            table_nodos = Table(data_nodos, colWidths=[1.0*inch]*len(df_nodos.columns))
            table_nodos.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), style_header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), style_header_font_color),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, style_data_color])
            ]))
            story.append(Paragraph("Tabla de Nodos", styles['Heading3']))
            story.append(table_nodos)
            story.append(PageBreak()) 
        
        # --- CORRECCIÓN: TABLA DE CONECTIVIDAD REINSERTADA ---
        df_conectividad = crear_tabla_conectividad()
        if not df_conectividad.empty:
            data_conectividad = [list(df_conectividad.columns)] + df_conectividad.values.tolist()
            col_widths_connect = [0.8*inch]*len(df_conectividad.columns)
            if sum(col_widths_connect) > doc.width:
                col_widths_connect = [(doc.width*0.95) / len(df_conectividad.columns)] * len(df_conectividad.columns)
                
            table_conectividad = Table(data_conectividad, colWidths=col_widths_connect)
            table_conectividad.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), style_header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), style_header_font_color),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, style_data_color])
            ]))
            story.append(Paragraph("Tabla de Conectividad", styles['Heading3']))
            story.append(table_conectividad)
            story.append(PageBreak())
        # --- FIN CORRECCIÓN ---

        # --- 4. MATRICES GLOBALES Y LOCALES (Una por página) ---
        story.append(Paragraph("4. MATRICES DEL SISTEMA", section_style))
        
        def create_matrix_table(matrix, title):
            # Asegurarse de que los GLs coincidan con la info (12 GLs en el ejemplo)
            gl_labels = [f"GL{info['numero']}" for info in st.session_state.grados_libertad_info]
            header = [''] + gl_labels
            
            data = [header]
            for i, row in enumerate(matrix):
                formatted_row = [gl_labels[i]] + [f'{val:.3e}' for val in row]
                data.append(formatted_row)
            
            num_cols = len(header)
            col_width = (doc.width*0.95) / num_cols
            if col_width < 0.3*inch: col_width = 0.3*inch # Mínimo
            
            t = Table(data, colWidths=[col_width]*num_cols)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), style_header_color),
                ('BACKGROUND', (0, 1), (0, -1), style_data_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), style_header_font_color),
                ('TEXTCOLOR', (0, 1), (-1, 0), style_header_font_color), # Color de GLs en 1ra col
                ('TEXTCOLOR', (1, 1), (-1, -1), style_data_font_color), # Color de datos
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 5),
                ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                ('BOX', (0, 0), (-1, -1), 1, colors.black)
            ]))
            return [Paragraph(title, styles['Heading3']), t, Spacer(1, 15)]

        story.extend(create_matrix_table(resultado['K_global'], "Matriz de Rigidez Global (K)"))
        story.append(PageBreak()) 
        story.extend(create_matrix_table(resultado['M_global'], "Matriz de Masa Global (M)"))
        story.append(PageBreak()) 

        story.append(Paragraph("Matrices Locales por Elemento", styles['Heading3']))

        if st.session_state.tipo_elemento == "viga_portico":
            labels = ["u1", "v1", "θ1", "u2", "v2", "θ2"]
        elif st.session_state.tipo_elemento == "viga":
            labels = ["v1", "θ1", "v2", "θ2"]
        else:
            labels = ["u1", "v1", "u2", "v2"]

        for elem_id, matrices in st.session_state.matrices_elementos.items():
            k_local = np.array(matrices['local'])
            m_local = np.array(matrices['masa_local'])
            
            data_k = [["K' Local"] + labels] + [[labels[i]] + [f"{v:.3e}" for v in row] for i, row in enumerate(k_local)]
            data_m = [["M' Local"] + labels] + [[labels[i]] + [f"{v:.3e}" for v in row] for i, row in enumerate(m_local)]
            
            col_width_local = (doc.width*0.9) / (len(labels)+1)
            
            t_k = Table(data_k, colWidths=[col_width_local]*(len(labels)+1))
            t_m = Table(data_m, colWidths=[col_width_local]*(len(labels)+1))
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), style_header_color),
                ('BACKGROUND', (0, 1), (0, -1), style_data_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), style_header_font_color),
                ('TEXTCOLOR', (0, 1), (0, -1), style_header_font_color), # 1ra col
                ('TEXTCOLOR', (1, 1), (-1, -1), style_data_font_color), # Datos
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 8), 
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black), 
                ('BOX', (0,0), (-1,-1), 1, colors.black)
            ])
            t_k.setStyle(style)
            t_m.setStyle(style)

            story.append(PageBreak()) 
            story.append(Paragraph(f"Elemento {elem_id} - Matriz de Rigidez Local (k')", styles['Heading3']))
            story.append(t_k)
            story.append(PageBreak()) 
            
            story.append(Paragraph(f"Elemento {elem_id} - Matriz de Masa Local (m')", styles['Heading3']))
            story.append(t_m)

        # Construir PDF
        doc.build(story)
        pdf_buffer.seek(0)
        return pdf_buffer
        
    except Exception as e:
        st.error(f"Error generando PDF: {str(e)}")
        return None

def generar_pdf_reporte_estatico():
    """Generar reporte PDF para análisis estático con estilo Excel (Corregido)"""
    if not st.session_state.resultados:
        return None
    
    try:
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4),
                                rightMargin=inch/2, leftMargin=inch/2,
                                topMargin=inch/2, bottomMargin=inch/2)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#2d3748'), spaceAfter=30, alignment=1)
        section_style = ParagraphStyle('CustomSection', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#4472C4'), spaceAfter=12, spaceBefore=12)
        
        # Esquema de color Azul (para TODAS las tablas)
        style_header_color = colors.HexColor('#4472C4')
        style_data_color = colors.HexColor('#D9E1F2')
        style_header_font_color = colors.whitesmoke
        style_data_font_color = colors.black
        
        story.append(Paragraph("ANÁLISIS ESTÁTICO - ESTRUCTURAS", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        resultado = st.session_state.resultados
        
        # 1. Métricas Principales
        story.append(Paragraph("1. RESUMEN DEL ANÁLISIS", section_style))
        metricas = [
            ("Tipo de Análisis:", "Estático"),
            ("Tipo de Elemento:", st.session_state.tipo_elemento.replace('_', ' ').title()),
            ("Número de Nodos:", len(st.session_state.nodos)),
            ("Número de Elementos:", len(st.session_state.elementos)),
            ("Total de DOF:", len(st.session_state.grados_libertad_info)),
            ("Determinante de K:", f"{resultado['determinante']:.6e}")
        ]
        data_metricas = []
        for label, value in metricas:
            data_metricas.append([Paragraph(label, styles['Normal']), Paragraph(str(value), styles['Normal'])])
        
        table_metricas = Table(data_metricas, colWidths=[2*inch, 4*inch])
        table_metricas.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(table_metricas)
        story.append(PageBreak())
        
        # 2. Tablas de Nodos y Conectividad (Una por página)
        story.append(Paragraph("2. INFORMACIÓN DE NODOS Y ELEMENTOS", section_style))
        
        df_nodos = crear_tabla_nodos()
        if not df_nodos.empty:
            data_nodos = [list(df_nodos.columns)] + df_nodos.values.tolist()
            table_nodos = Table(data_nodos, colWidths=[1.0*inch]*len(df_nodos.columns))
            table_nodos.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), style_header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), style_header_font_color),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, style_data_color])
            ]))
            story.append(Paragraph("Tabla de Nodos", styles['Heading3']))
            story.append(table_nodos)
            story.append(PageBreak()) 
        
        df_conectividad = crear_tabla_conectividad()
        if not df_conectividad.empty:
            data_conectividad = [list(df_conectividad.columns)] + df_conectividad.values.tolist()
            col_widths_connect = [0.8*inch]*len(df_conectividad.columns)
            if sum(col_widths_connect) > doc.width:
                col_widths_connect = [(doc.width*0.95) / len(df_conectividad.columns)] * len(df_conectividad.columns)
                
            table_conectividad = Table(data_conectividad, colWidths=col_widths_connect)
            table_conectividad.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), style_header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), style_header_font_color),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, style_data_color])
            ]))
            story.append(Paragraph("Tabla de Conectividad", styles['Heading3']))
            story.append(table_conectividad)
        
        story.append(PageBreak())
        
        # 3. Desplazamientos y Fuerzas (Una por página)
        story.append(Paragraph("3. DESPLAZAMIENTOS Y FUERZAS NODALES", section_style))
        
        # Desplazamientos
        df_desplazamientos = pd.DataFrame({
            'GL': [info['numero'] for info in st.session_state.grados_libertad_info],
            'Nodo': [info['nodo'] for info in st.session_state.grados_libertad_info],
            'Dirección': [info['direccion'] for info in st.session_state.grados_libertad_info],
            'Desplazamiento [m o rad]': [formatear_unidades(d, "desplazamiento") for d in resultado['desplazamientos'][np.array([info['numero']-1 for info in st.session_state.grados_libertad_info])]]
        })
        data_desplazamientos = [list(df_desplazamientos.columns)] + df_desplazamientos.values.tolist()
        table_desplazamientos = Table(data_desplazamientos, colWidths=[0.7*inch, 0.6*inch, 1*inch, 1.5*inch])
        table_desplazamientos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), style_header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), style_header_font_color),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, style_data_color])
        ]))
        story.append(Paragraph("Desplazamientos Nodales", styles['Heading3']))
        story.append(table_desplazamientos)
        story.append(PageBreak()) 
        
        # Fuerzas
        df_fuerzas = pd.DataFrame({
            'GL': [info['numero'] for info in st.session_state.grados_libertad_info],
            'Nodo': [info['nodo'] for info in st.session_state.grados_libertad_info],
            'Dirección': [info['direccion'] for info in st.session_state.grados_libertad_info],
            'Fuerza [N o Nm]': [formatear_unidades(f, "fuerza") for f in resultado['fuerzas'][np.array([info['numero']-1 for info in st.session_state.grados_libertad_info])]]
        })
        data_fuerzas = [list(df_fuerzas.columns)] + df_fuerzas.values.tolist()
        table_fuerzas = Table(data_fuerzas, colWidths=[0.7*inch, 0.6*inch, 1*inch, 1.5*inch])
        table_fuerzas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), style_header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), style_header_font_color),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, style_data_color])
        ]))
        story.append(Paragraph("Fuerzas en los Grados de Libertad", styles['Heading3']))
        story.append(table_fuerzas)
        
        doc.build(story)
        pdf_buffer.seek(0)
        return pdf_buffer
        
    except Exception as e:
        st.error(f"Error generando PDF: {str(e)}")
        return None

# --- Funciones de Reporte (Excel) ---

def generar_excel_reporte_dinamico():
    """Generar reporte Excel para análisis dinámico con múltiples hojas y estilo personalizado."""
    if not st.session_state.resultados_dinamicos:
        return None
    
    if not OPENPYXL_AVAILABLE:
        st.error("Se requiere instalar 'openpyxl' para exportar a Excel. Intente: pip install openpyxl")
        return None
    
    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        resultado = st.session_state.resultados_dinamicos
        
        # --- Definición de Estilos (basado en image_647baa.png) ---
        style_header_verde = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        style_header_gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        style_data_beige = PatternFill(start_color="FDF2CA", end_color="FDF2CA", fill_type="solid")
        style_header_azul = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Para matrices
        
        font_header = Font(bold=True, color="000000", size=11)
        font_header_blanco = Font(bold=True, color="FFFFFF", size=11)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def apply_style_to_range(ws, cell_range, style_fill=None, style_font=None, style_border=None, style_align=None):
            rows = ws[cell_range]
            if not isinstance(rows, tuple):
                rows = ((rows,),) # Asegurar que sea iterable

            for row in rows:
                for cell in row:
                    if style_fill: cell.fill = style_fill
                    if style_font: cell.font = style_font
                    if style_border: cell.border = style_border
                    if style_align: cell.alignment = style_align
        
        def auto_fit_cols(ws):
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: 
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # --- 1. HOJA: Modos de Vibración (Principal) ---
        ws_modos = wb.create_sheet("Modos de Vibración")
        df_modos = crear_tabla_modos_completa()
        
        # Escribir encabezados
        headers_modo = list(df_modos.columns)
        ws_modos.append(headers_modo)
        
        # Escribir datos
        for r_idx, row in enumerate(df_modos.values.tolist(), 2):
            ws_modos.append(row)
            
        # Aplicar Estilos
        num_rows = len(df_modos)
        num_cols = len(headers_modo)
        
        # Encabezados (Modo, w^2, Freq)
        apply_style_to_range(ws_modos, f'B1:{get_column_letter(num_cols)}3', style_header_verde, font_header, border, align_center)
        # Datos (Eigenvectors)
        apply_style_to_range(ws_modos, f'B4:{get_column_letter(num_cols)}{num_rows+1}', style_data_beige, None, border, align_center)
        # Columna de Parámetros (GLs)
        apply_style_to_range(ws_modos, f'A1:A{num_rows+1}', style_header_gris, font_header, border, align_left)
        # Formato de número
        for row in ws_modos[f'B2:{get_column_letter(num_cols)}{num_rows+1}']:
            for cell in row:
                if isinstance(cell.value, str) and ('e' in cell.value or '.' in cell.value):
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '0.00E+00'
                    except ValueError:
                        pass
        for row in ws_modos[f'B3:B3']: # Fila Freq
             for cell in row:
                try:
                    cell.value = float(cell.value)
                    cell.number_format = '0.0'
                except: pass
        
        auto_fit_cols(ws_modos)
        ws_modos.column_dimensions['A'].width = 25

        # --- 2. HOJA: Nodos ---
        ws_nodos = wb.create_sheet("Nodos")
        
        df_nodos = crear_tabla_nodos()
        if not df_nodos.empty:
            ws_nodos.append(list(df_nodos.columns))
            for row in df_nodos.values.tolist():
                ws_nodos.append(row)
            apply_style_to_range(ws_nodos, f'A1:{get_column_letter(len(df_nodos.columns))}1', style_header_azul, font_header_blanco, border, align_center)
            apply_style_to_range(ws_nodos, f'A2:{get_column_letter(len(df_nodos.columns))}{len(df_nodos)+1}', None, None, border, align_center)
            auto_fit_cols(ws_nodos)
        
        # --- 3. HOJA: ELEMENTOS ---
        ws_elem = wb.create_sheet("Elementos")
        
        df_elem = crear_tabla_conectividad()
        if not df_elem.empty:
            ws_elem.append(list(df_elem.columns))
            for row in df_elem.values.tolist():
                ws_elem.append(row)
            apply_style_to_range(ws_elem, f'A1:{get_column_letter(len(df_elem.columns))}1', style_header_azul, font_header_blanco, border, align_center)
            apply_style_to_range(ws_elem, f'A2:{get_column_letter(len(df_elem.columns))}{len(df_elem)+1}', None, None, border, align_center)
            auto_fit_cols(ws_elem)
        
        # --- 4 & 5. HOJAS: Matrices Globales K y M ---
        def write_matrix_sheet(wb, sheet_name, matrix, header_prefix):
            ws = wb.create_sheet(sheet_name)
            headers = [f'{header_prefix}{info["numero"]}' for info in st.session_state.grados_libertad_info]
            ws.append([sheet_name] + headers)
            
            for i, row_data in enumerate(matrix):
                info = st.session_state.grados_libertad_info[i]
                ws.append([f'{header_prefix}{info["numero"]}'] + row_data.tolist())
            
            # Estilos
            apply_style_to_range(ws, f'B1:{get_column_letter(len(headers)+1)}1', style_header_azul, font_header_blanco, border, align_center)
            apply_style_to_range(ws, f'A2:A{len(headers)+1}', style_header_azul, font_header_blanco, border, align_center)
            apply_style_to_range(ws, f'A1:A1', style_header_azul, font_header_blanco, border, align_center) # Esquina
            
            for row in ws[f'B2:{get_column_letter(len(headers)+1)}{len(headers)+1}']:
                for cell in row:
                    cell.number_format = '0.00E+00'
                    cell.border = border
            
            ws.column_dimensions['A'].width = 15
            for i in range(2, len(headers) + 2):
                ws.column_dimensions[get_column_letter(i)].width = 15
                
        write_matrix_sheet(wb, "Matriz Rigidez Global (K)", resultado['K_global'], "GL")
        write_matrix_sheet(wb, "Matriz Masa Global (M)", resultado['M_global'], "GL")

        # --- 6. HOJAS: Matrices Locales (K' y M') ---
        if st.session_state.tipo_elemento == "viga_portico":
            labels = ["u1", "v1", "θ1", "u2", "v2", "θ2"]
        elif st.session_state.tipo_elemento == "viga":
            labels = ["v1", "θ1", "v2", "θ2"]
        else:
            labels = ["u1", "v1", "u2", "v2"]
        
        for elem_id, matrices in st.session_state.matrices_elementos.items():
            k_local = np.array(matrices['local'])
            m_local = np.array(matrices['masa_local'])
            
            # Hoja para K'
            ws_k_loc = wb.create_sheet(f"Elem {elem_id} - K Local")
            ws_k_loc.append([f"Elem {elem_id} - K'"] + labels)
            for i, row_data in enumerate(k_local):
                ws_k_loc.append([labels[i]] + row_data.tolist())
            apply_style_to_range(ws_k_loc, f'B1:{get_column_letter(len(labels)+1)}1', style_header_azul, font_header_blanco, border, align_center)
            apply_style_to_range(ws_k_loc, f'A2:A{len(labels)+1}', style_header_azul, font_header_blanco, border, align_center)
            apply_style_to_range(ws_k_loc, f'A1:A1', style_header_azul, font_header_blanco, border, align_center)
            for row in ws_k_loc[f'B2:{get_column_letter(len(labels)+1)}{len(labels)+1}']:
                for cell in row:
                    cell.number_format = '0.00E+00'
                    cell.border = border
            auto_fit_cols(ws_k_loc)
            
            # Hoja para M'
            ws_m_loc = wb.create_sheet(f"Elem {elem_id} - M Local")
            ws_m_loc.append([f"Elem {elem_id} - M'"] + labels)
            for i, row_data in enumerate(m_local):
                ws_m_loc.append([labels[i]] + row_data.tolist())
            apply_style_to_range(ws_m_loc, f'B1:{get_column_letter(len(labels)+1)}1', style_header_azul, font_header_blanco, border, align_center)
            apply_style_to_range(ws_m_loc, f'A2:A{len(labels)+1}', style_header_azul, font_header_blanco, border, align_center)
            apply_style_to_range(ws_m_loc, f'A1:A1', style_header_azul, font_header_blanco, border, align_center)
            for row in ws_m_loc[f'B2:{get_column_letter(len(labels)+1)}{len(labels)+1}']:
                for cell in row:
                    cell.number_format = '0.00E+00'
                    cell.border = border
            auto_fit_cols(ws_m_loc)
            
        # Guardar en buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
        
    except Exception as e:
        st.error(f"Error generando Excel: {str(e)}")
        return None

def generar_excel_reporte_estatico():
    """Generar reporte Excel para análisis estático con múltiples hojas"""
    if not st.session_state.resultados:
        return None
    
    if not OPENPYXL_AVAILABLE:
        st.error("Se requiere instalar 'openpyxl' para exportar a Excel. Intente: pip install openpyxl")
        return None
    
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        resultado = st.session_state.resultados
        
        # Estilos
        header_style = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def apply_style_to_range(ws, cell_range, style_fill=None, style_font=None, style_border=None, style_align=None):
            rows = ws[cell_range]
            if not isinstance(rows, tuple):
                rows = ((rows,),) # Asegurar que sea iterable

            for row in rows:
                for cell in row:
                    if style_fill: cell.fill = style_fill
                    if style_font: cell.font = style_font
                    if style_border: cell.border = style_border
                    if style_align: cell.alignment = style_align
        
        def auto_fit_cols(ws):
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: 
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
        
        # 1. HOJA: RESUMEN
        ws_resumen = wb.create_sheet("Resumen")
        
        cell_titulo = ws_resumen.cell(row=1, column=1, value="ANÁLISIS ESTÁTICO - RESUMEN")
        cell_titulo.font = Font(bold=True, size=14, color="FFFFFF")
        cell_titulo.fill = header_fill
        ws_resumen.merge_cells('A1:D1')
        
        row = 3
        info_general = [
            ("Tipo de Análisis:", "Estático - Análisis de Cargas"),
            ("Tipo de Elemento:", st.session_state.tipo_elemento.replace('_', ' ').title()),
            ("Número de Nodos:", len(st.session_state.nodos)),
            ("Número de Elementos:", len(st.session_state.elementos)),
            ("Total de DOF:", len(st.session_state.grados_libertad_info)),
            ("Determinante de K:", f"{resultado['determinante']:.6e}"),
            ("Fecha de Generación:", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        ]
        
        for label, value in info_general:
            cell_label = ws_resumen.cell(row=row, column=1, value=label)
            cell_label.font = Font(bold=True)
            ws_resumen.cell(row=row, column=2, value=str(value))
            row += 1
        
        # 2. HOJA: NODOS
        ws_nodos = wb.create_sheet("Nodos")
        
        df_nodos = crear_tabla_nodos()
        if not df_nodos.empty:
            ws_nodos.append(list(df_nodos.columns))
            for row in df_nodos.values.tolist():
                ws_nodos.append(row)
            apply_style_to_range(ws_nodos, f'A1:{get_column_letter(len(df_nodos.columns))}1', header_fill, header_style, border, header_alignment)
            apply_style_to_range(ws_nodos, f'A2:{get_column_letter(len(df_nodos.columns))}{len(df_nodos)+1}', None, None, border, header_alignment)
            auto_fit_cols(ws_nodos)
        
        # 3. HOJA: ELEMENTOS
        ws_elem = wb.create_sheet("Elementos")
        
        df_elem = crear_tabla_conectividad()
        if not df_elem.empty:
            ws_elem.append(list(df_elem.columns))
            for row in df_elem.values.tolist():
                ws_elem.append(row)
            apply_style_to_range(ws_elem, f'A1:{get_column_letter(len(df_elem.columns))}1', header_fill, header_style, border, header_alignment)
            apply_style_to_range(ws_elem, f'A2:{get_column_letter(len(df_elem.columns))}{len(df_elem)+1}', None, None, border, header_alignment)
            auto_fit_cols(ws_elem)
        
        # 4. HOJA: DESPLAZAMIENTOS
        ws_desp = wb.create_sheet("Desplazamientos")
        
        headers_desp = ["GL", "Nodo", "Dirección", "Desplazamiento [m o rad]"]
        ws_desp.append(headers_desp)
        apply_style_to_range(ws_desp, 'A1:D1', header_fill, header_style, border, header_alignment)
        
        for row_idx, info in enumerate(st.session_state.grados_libertad_info, 2):
            gl_num = info['numero']
            despl = resultado['desplazamientos'][gl_num - 1]
            
            ws_desp.cell(row=row_idx, column=1, value=f"GL{gl_num}").border = border
            ws_desp.cell(row=row_idx, column=2, value=info['nodo']).border = border
            ws_desp.cell(row=row_idx, column=3, value=info['direccion']).border = border
            
            cell_desp = ws_desp.cell(row=row_idx, column=4, value=despl)
            cell_desp.number_format = '0.000000E+00'
            cell_desp.border = border
        
        auto_fit_cols(ws_desp)
        
        # 5. HOJA: FUERZAS
        ws_fuerzas = wb.create_sheet("Fuerzas")
        
        headers_fuerzas = ["GL", "Nodo", "Dirección", "Fuerza [N o Nm]"]
        ws_fuerzas.append(headers_fuerzas)
        apply_style_to_range(ws_fuerzas, 'A1:D1', header_fill, header_style, border, header_alignment)
        
        for row_idx, info in enumerate(st.session_state.grados_libertad_info, 2):
            gl_num = info['numero']
            fuerza = resultado['fuerzas'][gl_num - 1]
            
            ws_fuerzas.cell(row=row_idx, column=1, value=f"GL{gl_num}").border = border
            ws_fuerzas.cell(row=row_idx, column=2, value=info['nodo']).border = border
            ws_fuerzas.cell(row=row_idx, column=3, value=info['direccion']).border = border
            
            cell_fuerza = ws_fuerzas.cell(row=row_idx, column=4, value=fuerza)
            cell_fuerza.number_format = '0.000000E+00'
            cell_fuerza.border = border
        
        auto_fit_cols(ws_fuerzas)
        
        # 6. HOJA: MATRIZ RIGIDEZ
        ws_k = wb.create_sheet("Matriz Rigidez")
        
        K_global = resultado['K_global']
        
        headers_k = [f"GL{i+1}" for i in range(K_global.shape[1])]
        ws_k.append(["Matriz K"] + headers_k)
        
        for i, row_data in enumerate(K_global):
            ws_k.append([f"GL{i+1}"] + row_data.tolist())
        
        apply_style_to_range(ws_k, f'B1:{get_column_letter(len(headers_k)+1)}1', header_fill, header_style, border, header_alignment)
        apply_style_to_range(ws_k, f'A2:A{len(headers_k)+1}', header_fill, header_style, border, header_alignment)
        apply_style_to_range(ws_k, 'A1:A1', header_fill, header_style, border, header_alignment)
        
        for row in ws_k[f'B2:{get_column_letter(len(headers_k)+1)}{len(headers_k)+1}']:
            for cell in row:
                cell.number_format = '0.00E+00'
                cell.border = border
        
        ws_k.column_dimensions['A'].width = 15
        for i in range(2, len(headers_k) + 2):
            ws_k.column_dimensions[get_column_letter(i)].width = 15
        
        # Guardar en buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
        
    except Exception as e:
        st.error(f"Error generando Excel: {str(e)}")
        return None

# -----------------------------------------------------------------
# 5. INICIALIZACIÓN DE SESSION STATE
# -----------------------------------------------------------------

if 'step' not in st.session_state:
    st.session_state.step = 0
if 'tipo_elemento' not in st.session_state:
    st.session_state.tipo_elemento = None
if 'modo' not in st.session_state:
    st.session_state.modo = None
if 'tipo_analisis' not in st.session_state:
    st.session_state.tipo_analisis = None  # 'estatico' o 'dinamico'
if 'usuario_nombre' not in st.session_state:
    st.session_state.usuario_nombre = ""
if 'nodos' not in st.session_state:
    st.session_state.nodos = []
if 'elementos' not in st.session_state:
    st.session_state.elementos = []
if 'matrices_elementos' not in st.session_state:
    st.session_state.matrices_elementos = {}
if 'grados_libertad_info' not in st.session_state:
    st.session_state.grados_libertad_info = []
if 'nombres_fuerzas' not in st.session_state:
    st.session_state.nombres_fuerzas = {}
if 'resultados' not in st.session_state:
    st.session_state.resultados = None
if 'materiales_personalizados' not in st.session_state:
    st.session_state.materiales_personalizados = {}
if 'auto_calcular' not in st.session_state:
    st.session_state.auto_calcular = True
if 'nodo_seleccionado_interactivo' not in st.session_state:
    st.session_state.nodo_seleccionado_interactivo = None
if 'nodos_interactivos' not in st.session_state:
    st.session_state.nodos_interactivos = []
if 'elementos_interactivos' not in st.session_state:
    st.session_state.elementos_interactivos = []
if 'num_nodos' not in st.session_state:
    st.session_state.num_nodos = 2
if 'num_fijos' not in st.session_state:
    st.session_state.num_fijos = 1
if 'num_elementos' not in st.session_state:
    st.session_state.num_elementos = 1
if 'condiciones_contorno_dinamica' not in st.session_state:
    st.session_state.condiciones_contorno_dinamica = {}
if 'modo_visualizacion' not in st.session_state:
    st.session_state.modo_visualizacion = 1
if 'resultados_dinamicos' not in st.session_state:
    st.session_state.resultados_dinamicos = None
if 'grupos_elementos' not in st.session_state: # (Añadido de V4.7)
    st.session_state.grupos_elementos = {}


# -----------------------------------------------------------------
# 6. NAVEGACIÓN DE LA APLICACIÓN (LÓGICA DE PÁGINA)
# -----------------------------------------------------------------

# Llamar a las funciones de UI (definidas en la Parte 1)
mostrar_barra_progreso()
mostrar_sidebar_mejorado()

if st.session_state.step == 0:
    st.markdown("""
    <div style='background: linear-gradient(135deg, #2d3748 0%, #4a5568 100%); 
                min-height: 80vh; display: flex; align-items: center; justify-content: center; 
                margin: -1rem; padding: 2rem; border-radius: 15px;'>
        <div style='max-width: 1200px; text-align: center;'>
            <h1 style='font-size: 4rem; margin-bottom: 1rem; color: white; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);'>
                Análisis Estructural
            </h1>
            <p style='font-size: 1.5rem; color: rgba(255,255,255,0.9); margin-bottom: 3rem; text-shadow: 1px 1px 2px rgba(0,0,0,0.3);'>
                Método de Matrices • Seleccione el Tipo de Análisis
            </p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("## Seleccione el Tipo de Análisis")
    
    col1, col2 = st.columns(2, gap="large")
    
    with col1:
        st.markdown("""
        <div style='background: white; padding: 2rem; border-radius: 15px; border: 2px solid #e2e8f0; 
                    box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin: 1rem 0;'>
            <h3 style='color: #1a202c; font-size: 1.8rem; margin-bottom: 1rem; text-align: center;'>📊 Análisis Estático</h3>
            <p style='color: #4a5568; line-height: 1.6; margin-bottom: 1.5rem; text-align: center;'>
                Análisis bajo cargas estáticas<br><br>
                • Cálculo de desplazamientos<br>
                • Matriz de rigidez global<br>
                • Fuerzas internas
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("ANÁLISIS ESTÁTICO", key="estatico_analysis", type="primary", use_container_width=True):
            set_tipo_analisis("estatico")
    
    with col2:
        st.markdown("""
        <div style='background: white; padding: 2rem; border-radius: 15px; border: 2px solid #e2e8f0; 
                    box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin: 1rem 0;'>
            <h3 style='color: #1a202c; font-size: 1.8rem; margin-bottom: 1rem; text-align: center;'>🎵 Análisis Dinámico</h3>
            <p style='color: #4a5568; line-height: 1.6; margin-bottom: 1.5rem; text-align: center;'>
                Análisis de vibraciones libres<br><br>
                • Frecuencias naturales<br>
                • Matriz de masa global<br>
                • Modos de vibración
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("ANÁLISIS DINÁMICO", key="dinamico_analysis", type="primary", use_container_width=True):
            set_tipo_analisis("dinamico")

elif st.session_state.step == 1:
    st.markdown("## Seleccione el Tipo de Elemento Estructural")
    
    col1, col2, col3 = st.columns(3, gap="large")
    
    with col1:
        st.markdown("""
        <div style='background: white; padding: 2rem; border-radius: 15px; border: 2px solid #e2e8f0; 
                    box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin: 1rem 0;'>
            <h3 style='color: #1a202c; font-size: 1.8rem; margin-bottom: 1rem; text-align: center;'>🔗 Barra</h3>
            <p style='color: #4a5568; line-height: 1.6; margin-bottom: 1.5rem; text-align: center;'>
                Elementos que solo resisten fuerzas axiales.<br><br>
                • 2 grados de libertad por nodo (X, Y)<br>
                • Matriz de rigidez 4×4
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("SELECCIONAR BARRA", key="barra_mode", type="primary", use_container_width=True):
            set_tipo_elemento("barra")
    
    with col2:
        st.markdown("""
        <div style='background: white; padding: 2rem; border-radius: 15px; border: 2px solid #e2e8f0; 
                    box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin: 1rem 0;'>
            <h3 style='color: #1a202c; font-size: 1.8rem; margin-bottom: 1rem; text-align: center;'>📏 Viga</h3>
            <p style='color: #4a5568; line-height: 1.6; margin-bottom: 1.5rem; text-align: center;'>
                Elementos que resisten flexión pura.<br><br>
                • 2 grados de libertad por nodo (Y, θ)<br>
                • Matriz de rigidez 4×4
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("SELECCIONAR VIGA", key="viga_mode", type="primary", use_container_width=True):
            set_tipo_elemento("viga")
    
    with col3:
        st.markdown("""
        <div style='background: white; padding: 2rem; border-radius: 15px; border: 2px solid #e2e8f0; 
                    box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin: 1rem 0;'>
            <h3 style='color: #1a202c; font-size: 1.8rem; margin-bottom: 1rem; text-align: center;'>🏗️ Viga Pórtico</h3>
            <p style='color: #4a5568; line-height: 1.6; margin-bottom: 1.5rem; text-align: center;'>
                Elementos que combinan barra y viga.<br><br>
                • 3 grados de libertad por nodo (X, Y, θ)<br>
                • Matriz de rigidez 6×6
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("SELECCIONAR VIGA PÓRTICO", key="viga_portico_mode", type="primary", use_container_width=True):
            set_tipo_elemento("viga_portico")

elif st.session_state.step == 2:
    st.markdown("## Información del Usuario")
    usuario_nombre = st.text_input("Nombre del usuario", value=st.session_state.usuario_nombre)
    
    if st.button("Continuar →", type="primary"):
        st.session_state.usuario_nombre = usuario_nombre
        next_step()

elif st.session_state.step == 3:
    st.markdown("## Selección de Modo")
    st.markdown(f"Seleccione el modo de trabajo para el análisis")
    
    col1, col2 = st.columns(2, gap="large")
    
    with col1:
        st.markdown("""
        <div style='background: white; padding: 2rem; border-radius: 15px; border: 2px solid #e2e8f0; 
                    box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin: 1rem 0;'>
            <h3 style='color: #1a202c; font-size: 1.5rem; margin-bottom: 1rem; text-align: center;'>Manual</h3>
            <p style='color: #4a5568; line-height: 1.6; text-align: center;'>
                Ingrese coordenadas exactas de nodos
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("SELECCIONAR MANUAL", key="manual_mode", type="primary", use_container_width=True):
            set_modo("manual")
    
    with col2:
        st.markdown("""
        <div style='background: white; padding: 2rem; border-radius: 15px; border: 2px solid #e2e8f0; 
                    box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin: 1rem 0;'>
            <h3 style='color: #1a202c; font-size: 1.5rem; margin-bottom: 1rem; text-align: center;'>Interactivo</h3>
            <p style='color: #4a5568; line-height: 1.6; text-align: center;'>
                Diseño visual e intuitivo
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("SELECCIONAR INTERACTIVO", key="interactive_mode", type="primary", use_container_width=True):
            set_modo("interactivo")

# --- LÓGICA MODO MANUAL ---
elif st.session_state.step == 4 and st.session_state.modo == "manual":
    st.markdown("## Número de Nodos")
    st.markdown("Ingrese la cantidad total de nodos en su estructura.")
    num_nodos = st.number_input("Cantidad de Nodos", min_value=2, value=st.session_state.num_nodos, step=1)
    
    if st.button("Continuar →", type="primary"):
        st.session_state.num_nodos = num_nodos
        # Resetear nodos y elementos si el número cambia
        st.session_state.nodos = []
        st.session_state.elementos = []
        next_step()

elif st.session_state.step == 5 and st.session_state.modo == "manual":
    st.markdown("## Coordenadas de Nodos")
    st.markdown("Ingrese las coordenadas X e Y para cada nodo.")
    
    nodos_actualizados = []
    if st.session_state.modo == "manual":
        for i in range(st.session_state.num_nodos):
            nodo_id = i + 1
            
            # Obtener datos previos si existen
            nodo_existente = next((n for n in st.session_state.nodos if n['id'] == nodo_id), {})
            x = nodo_existente.get('x', 0.0)
            y = nodo_existente.get('y', 0.0)
            
            col1, col2 = st.columns(2)
            with col1:
                x_coord = st.number_input(f"Nodo {nodo_id} - X [m]", value=x, key=f"nodo_x_{nodo_id}")
            with col2:
                y_coord = st.number_input(f"Nodo {nodo_id} - Y [m]", value=y, key=f"nodo_y_{nodo_id}")

            # Preservar el diccionario existente (como 'tipo') y solo actualizar coords
            nodo_actualizado = nodo_existente.copy()
            nodo_actualizado['id'] = nodo_id
            nodo_actualizado['x'] = x_coord
            nodo_actualizado['y'] = y_coord
            nodos_actualizados.append(nodo_actualizado)
            
        st.session_state.nodos = nodos_actualizados
    
    if st.button("Continuar →", type="primary"):
        next_step()

elif st.session_state.step == 6 and st.session_state.modo == "manual":
    st.markdown("## Clasificación de Nodos")
    st.markdown("Defina el tipo de cada nodo (Fijo o Libre).")
    
    if len(st.session_state.nodos) != st.session_state.num_nodos:
        st.warning(f"Por favor, defina las coordenadas de los {st.session_state.num_nodos} nodos primero.")
        st.stop()
        
    nodos_clasificados = []
    for i in range(st.session_state.num_nodos):
        nodo_id = i + 1
        nodo_actual = next((n for n in st.session_state.nodos if n['id'] == nodo_id), None)
        
        if not nodo_actual:
            st.error(f"Datos del Nodo {nodo_id} incompletos. Por favor, revise los pasos anteriores.")
            st.stop()

        tipo = st.radio(
            f"Nodo {nodo_id} ({nodo_actual['x']:.2f}, {nodo_actual['y']:.2f})",
            ("Libre", "Fijo"),
            index=0 if nodo_actual.get('tipo', 'Libre') == 'Libre' else 1,
            key=f"tipo_nodo_{nodo_id}"
        )
        nodo_actual['tipo'] = tipo.lower()
        nodos_clasificados.append(nodo_actual)

    if st.button("Continuar →", type="primary"):
        st.session_state.nodos = nodos_clasificados
        next_step()

elif st.session_state.step == 7 and st.session_state.modo == "manual":
    st.markdown("## Número de Elementos")
    st.markdown("Ingrese la cantidad total de elementos estructurales.")
    num_elementos = st.number_input("Cantidad de Elementos", min_value=1, value=st.session_state.num_elementos, step=1)
    
    if st.button("Continuar →", type="primary"):
        st.session_state.num_elementos = num_elementos
        # Resetear elementos si el número cambia
        st.session_state.elementos = []
        st.session_state.matrices_elementos = {}
        next_step()

# --- LÓGICA MODO INTERACTIVO ---
elif st.session_state.step == 4 and st.session_state.modo == "interactivo":
    st.markdown("## Editor Interactivo de Estructura")
    st.markdown(f"Utilice el gráfico para crear su estructura con elementos tipo **{st.session_state.tipo_elemento.replace('_', ' ').title()}**. Haga clic para añadir nodos y conectarlos para formar elementos.")
    
    # Columnas para controles
    col1, col2, col3 = st.columns(3)
    
    with col1:
        tipo_nodo = st.radio("Tipo de nodo a añadir:", ["libre", "fijo"])
    
    with col2:
        if st.button("🗑️ Limpiar Todo", type="secondary"):
            st.session_state.nodos_interactivos = []
            st.session_state.elementos_interactivos = []
            st.session_state.nodo_seleccionado_interactivo = None
            st.rerun()
    
    with col3:
        if st.button("✅ Finalizar Diseño y Continuar", type="primary"):
            if len(st.session_state.nodos_interactivos) >= 2 and len(st.session_state.elementos_interactivos) >= 1:
                transferir_datos_interactivos() # Esta función lleva al paso 8
            else:
                st.error("Necesita al menos 2 nodos y 1 elemento para continuar")
    
    # Crear gráfico interactivo moderno
    fig = crear_grafico_interactivo_moderno()
    st.plotly_chart(fig, use_container_width=True)

    # Controles para añadir nodos manualmente
    st.markdown("### Añadir Nodos Manualmente")
    col1, col2, col3 = st.columns(3)

    with col1:
        x_nuevo = st.number_input("Coordenada X", value=0.0, format="%.2f", key="x_nuevo")

    with col2:
        y_nuevo = st.number_input("Coordenada Y", value=0.0, format="%.2f", key="y_nuevo")

    with col3:
        st.write("Añadir nodo con coordenadas:")
        if st.button("➕ Añadir Nodo", type="primary", use_container_width=True):
            nodo_id = agregar_nodo_interactivo(x_nuevo, y_nuevo, tipo_nodo)
            st.success(f"Nodo {nodo_id} ({tipo_nodo}) añadido en ({x_nuevo:.2f}, {y_nuevo:.2f})")
            st.rerun()

    # Controles para crear elementos
    if len(st.session_state.nodos_interactivos) >= 2:
        st.markdown(f"### Crear {st.session_state.tipo_elemento.replace('_', ' ').title()}s")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            nodos_disponibles = [n['id'] for n in st.session_state.nodos_interactivos]
            nodo_inicio_sel = st.selectbox("Nodo Inicio", nodos_disponibles, key="nodo_inicio_sel")
        
        with col2:
            nodos_fin_disponibles = [n for n in nodos_disponibles if n != nodo_inicio_sel]
            nodo_fin_sel = st.selectbox("Nodo Fin", nodos_fin_disponibles, key="nodo_fin_sel")
        
        with col3:
            st.write("Conectar nodos:")
            if st.button(f"🔗 Crear {st.session_state.tipo_elemento.title()}", type="primary", use_container_width=True):
                if not nodos_fin_disponibles:
                    st.error("Se necesitan al menos dos nodos para crear un elemento.")
                else:
                    elemento_id = agregar_elemento_interactivo(nodo_inicio_sel, nodo_fin_sel)
                    if elemento_id:
                        st.success(f"{st.session_state.tipo_elemento.title()} {elemento_id} creado entre nodos {nodo_inicio_sel} y {nodo_fin_sel}")
                        st.rerun()
                    else:
                        st.warning("No se pudo crear el elemento (puede que ya exista)")
    
    # Mostrar nodos y elementos en tablas
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### Nodos")
        if st.session_state.nodos_interactivos:
            for nodo in st.session_state.nodos_interactivos:
                col_a, col_b = st.columns([3, 1])
                with col_a:
                    st.markdown(f"Nodo {nodo['id']} ({nodo['tipo']}): ({nodo['x']:.2f}, {nodo['y']:.2f})")
                with col_b:
                    if st.button(f"🗑️", key=f"del_nodo_{nodo['id']}"):
                        eliminar_nodo_interactivo(nodo['id'])
    
    with col2:
        st.markdown(f"### {st.session_state.tipo_elemento.replace('_', ' ').title()}s")
        if st.session_state.elementos_interactivos:
            for elem in st.session_state.elementos_interactivos:
                col_a, col_b = st.columns([3, 1])
                with col_a:
                    st.markdown(f"{st.session_state.tipo_elemento.title()} {elem['id']}: Nodo {elem['nodo_inicio']} → Nodo {elem['nodo_fin']}")
                with col_b:
                    if st.button(f"🗑️", key=f"del_elem_{elem['id']}"):
                        eliminar_elemento_interactivo(elem['id'])

# --- PASOS UNIFICADOS (Manual e Interactivo) ---

elif st.session_state.step == 8:
    st.markdown("## Definición de Elementos")
    st.markdown(f"Configure las propiedades de cada elemento tipo **{st.session_state.tipo_elemento.replace('_', ' ').title()}**.")
    
    # (Lógica de V4.7 para Agrupación de Elementos)
    with st.expander("👥 Agrupación de Elementos", expanded=False):
        st.markdown("Agrupe elementos con propiedades similares para configuración masiva")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            nombre_grupo = st.text_input("Nombre del grupo:", placeholder="Ej: Columnas, Vigas principales")
        with col2:
            elementos_disponibles = [f"Elemento {e['id']}" for e in st.session_state.elementos]
            elementos_seleccionados = st.multiselect("Seleccionar elementos:", elementos_disponibles)
        with col3:
            st.write("Crear nuevo grupo:")
            if st.button("➕ Crear Grupo") and nombre_grupo and elementos_seleccionados:
                if nombre_grupo not in st.session_state.grupos_elementos:
                    st.session_state.grupos_elementos[nombre_grupo] = {
                        'elementos': [int(elem.split()[-1]) for elem in elementos_seleccionados],
                        'material': None,
                        'tipo_seccion': None,
                        'parametros_seccion': {}
                    }
                    st.success(f"Grupo '{nombre_grupo}' creado con {len(elementos_seleccionados)} elementos")
                    st.rerun()
                else:
                    st.warning("Ya existe un grupo con ese nombre")
        
        # Lógica para configurar grupos existentes
        if st.session_state.grupos_elementos:
            st.markdown("#### Grupos Existentes")
            for nombre_grupo, info_grupo in st.session_state.grupos_elementos.items():
                with st.container():
                    col1, col2, col3 = st.columns([2, 2, 1])
                    with col1:
                        st.markdown(f"**{nombre_grupo}**")
                        st.caption(f"Elementos: {info_grupo['elementos']}")
                    with col3:
                        if st.button(f"🗑️", key=f"del_grupo_{nombre_grupo}"):
                            del st.session_state.grupos_elementos[nombre_grupo]
                            st.rerun()
                    
                    # Configuración del grupo
                    st.markdown(f"##### Configuración para grupo '{nombre_grupo}'")
                    
                    todos_materiales = {**MATERIALES_AEROESPACIALES, **st.session_state.materiales_personalizados}
                    nombres_materiales = list(todos_materiales.keys())
                    material_grupo = st.selectbox("Material del grupo:", nombres_materiales, key=f"mat_grupo_{nombre_grupo}")
                    
                    tipo_seccion_grupo = st.radio("Tipo de sección:", 
                                                ["circular_solida", "circular_hueca", "rectangular", "cuadrada"], 
                                                format_func=lambda x: x.replace('_', ' ').title(), 
                                                key=f"seccion_grupo_{nombre_grupo}")
                    
                    parametros_grupo = {}
                    if tipo_seccion_grupo == "circular_solida":
                        radio_grupo = st.number_input("Radio (m):", value=0.01, min_value=0.001, format="%.4f", key=f"radio_grupo_{nombre_grupo}")
                        parametros_grupo['radio'] = radio_grupo
                    elif tipo_seccion_grupo == "circular_hueca":
                        radio_ext_grupo = st.number_input("Radio Exterior (m):", value=0.02, min_value=0.001, format="%.4f", key=f"radio_ext_grupo_{nombre_grupo}")
                        radio_int_grupo = st.number_input("Radio Interior (m):", value=0.01, min_value=0.0, max_value=radio_ext_grupo*0.99, format="%.4f", key=f"radio_int_grupo_{nombre_grupo}")
                        parametros_grupo['radio_ext'] = radio_ext_grupo
                        parametros_grupo['radio_int'] = radio_int_grupo
                    elif tipo_seccion_grupo == "rectangular":
                        lado1_grupo = st.number_input("Base (m):", value=0.02, min_value=0.001, format="%.4f", key=f"lado1_grupo_{nombre_grupo}")
                        lado2_grupo = st.number_input("Altura (m):", value=0.01, min_value=0.001, format="%.4f", key=f"lado2_grupo_{nombre_grupo}")
                        parametros_grupo['lado1'] = lado1_grupo
                        parametros_grupo['lado2'] = lado2_grupo
                    elif tipo_seccion_grupo == "cuadrada":
                        lado_grupo = st.number_input("Lado (m):", value=0.02, min_value=0.001, format="%.4f", key=f"lado_grupo_{nombre_grupo}")
                        parametros_grupo['lado'] = lado_grupo
                    
                    # Densidad para análisis dinámico
                    if st.session_state.tipo_analisis == "dinamico":
                        default_densidad = todos_materiales[material_grupo].get('densidad', 2700)
                        densidad_grupo = st.number_input("Densidad (kg/m³):", value=float(default_densidad), min_value=0.0, format="%.2f", key=f"densidad_grupo_{nombre_grupo}")
                    
                    if st.button(f"💾 Aplicar a Grupo '{nombre_grupo}'", key=f"aplicar_grupo_{nombre_grupo}"):
                        elementos_grupo_ids = info_grupo['elementos']
                        props_material = todos_materiales[material_grupo]
                        
                        for elemento_id in elementos_grupo_ids:
                            elemento_idx = next((i for i, e in enumerate(st.session_state.elementos) if e['id'] == elemento_id), None)
                            if elemento_idx is None:
                                continue # Saltar si el elemento no existe

                            # Actualizar propiedades del elemento
                            elem = st.session_state.elementos[elemento_idx]
                            elem['material'] = material_grupo
                            elem['tipo_seccion'] = tipo_seccion_grupo
                            elem['parametros_seccion'] = parametros_grupo
                            
                            area_final = calcular_area_seccion(tipo_seccion_grupo, parametros_grupo)
                            inercia_final = calcular_momento_inercia(tipo_seccion_grupo, parametros_grupo) if st.session_state.tipo_elemento in ["viga", "viga_portico"] else 0
                            
                            elem['area'] = area_final
                            elem['inercia'] = inercia_final
                            
                            if st.session_state.tipo_analisis == "dinamico":
                                elem['densidad'] = densidad_grupo
                            
                            # Recalcular matrices
                            nodo_inicio_obj = next(n for n in st.session_state.nodos if n['id'] == elem['nodo_inicio'])
                            nodo_fin_obj = next(n for n in st.session_state.nodos if n['id'] == elem['nodo_fin'])
                            
                            elem['longitud'] = calcular_longitud_elemento(nodo_inicio_obj, nodo_fin_obj)
                            elem['beta'] = calcular_angulo_beta(nodo_inicio_obj, nodo_fin_obj)
                            
                            E = props_material['modulo_young']
                            A, I, L, beta = elem['area'], elem['inercia'], elem['longitud'], elem['beta']
                            
                            k_global, k_local, m_global, m_local = None, None, None, None
                            
                            if st.session_state.tipo_elemento == "barra":
                                k_global, k_local = generar_matriz_rigidez_barra(E, A, L, beta)
                                if st.session_state.tipo_analisis == "dinamico":
                                    m_local = generar_matriz_masa_barra(elem['densidad'], A, L)
                                    # (Falta T para masa de barra)
                                    m_global = m_local # Placeholder
                            elif st.session_state.tipo_elemento == "viga":
                                k_global, k_local = generar_matriz_rigidez_viga(E, I, L)
                                if st.session_state.tipo_analisis == "dinamico":
                                    m_local = generar_matriz_masa_viga(elem['densidad'], A, L)
                                    m_global = m_local
                            else: # viga_portico
                                k_global, k_local = generar_matriz_rigidez_viga_portico(E, A, I, L, beta)
                                if st.session_state.tipo_analisis == "dinamico":
                                    m_local = generar_matriz_masa_viga_portico(elem['densidad'], A, L)
                                    T = generar_matriz_transformacion_viga_portico(beta)
                                    m_global = T.T @ m_local @ T

                            st.session_state.matrices_elementos[elemento_id] = {
                                'numerica': k_global.tolist() if k_global is not None else [],
                                'local': k_local.tolist() if k_local is not None else [],
                                'masa_global': m_global.tolist() if m_global is not None else [],
                                'masa_local': m_local.tolist() if m_local is not None else []
                            }
                        
                        st.success(f"✅ Configuración aplicada a {len(elementos_grupo_ids)} elementos del grupo '{nombre_grupo}'")
                        st.rerun()
                    st.divider()

    st.markdown("### Configuración Individual (Anula la configuración de grupo)")
    
    elementos_configurados = True
    # Asegurarse de que st.session_state.elementos tenga el número correcto de elementos
    # (El modo manual va a 7, luego 8. El interactivo va de 4 a 8)
    if len(st.session_state.elementos) < st.session_state.num_elementos:
        # Pre-poblar elementos si vienen del modo manual
        for i in range(st.session_state.num_elementos):
            if not any(e['id'] == i+1 for e in st.session_state.elementos):
                st.session_state.elementos.append({
                    'id': i+1,
                    'nodo_inicio': 1,
                    'nodo_fin': 2
                })
        # Ordenar por ID
        st.session_state.elementos.sort(key=lambda x: x['id'])


    for i in range(st.session_state.num_elementos):
        elemento_id = i + 1
        
        # Obtener datos previos (pueden venir del modo interactivo o de un grupo)
        elemento_existente = st.session_state.elementos[i]
        
        with st.expander(f"🔧 Elemento {elemento_id} (Nodos {elemento_existente['nodo_inicio']} → {elemento_existente['nodo_fin']})", expanded=False):
            
            default_material = elemento_existente.get('material') or list(MATERIALES_AEROESPACIALES.keys())[0]
            default_tipo_seccion = elemento_existente.get('tipo_seccion') or 'circular_solida'
            default_params = elemento_existente.get('parametros_seccion') or {}
            
            todos_materiales = {**MATERIALES_AEROESPACIALES, **st.session_state.materiales_personalizados}
            nombres_materiales = list(todos_materiales.keys())
            material_idx = nombres_materiales.index(default_material) if default_material in nombres_materiales else 0
            
            material_seleccionado = st.selectbox(f"Material", nombres_materiales, index=material_idx, key=f"material_ind_{elemento_id}")
            props_material = todos_materiales[material_seleccionado]
            
            st.markdown(f"E = {formatear_unidades(props_material['modulo_young'], 'presion')}")
            
            tipo_seccion = st.radio("Tipo de sección:", 
                                    ["circular_solida", "circular_hueca", "rectangular", "cuadrada"], 
                                    format_func=lambda x: x.replace('_', ' ').title(), 
                                    index=["circular_solida", "circular_hueca", "rectangular", "cuadrada"].index(default_tipo_seccion),
                                    key=f"tipo_seccion_ind_{elemento_id}")
            
            parametros_seccion = {}
            if tipo_seccion == "circular_solida":
                radio = st.number_input(f"Radio (m)", value=default_params.get('radio', 0.01), min_value=0.001, format="%.4f", key=f"radio_ind_{elemento_id}")
                parametros_seccion['radio'] = radio
            elif tipo_seccion == "circular_hueca":
                radio_ext = st.number_input(f"Radio Exterior (m):", value=default_params.get('radio_ext', 0.02), min_value=0.001, format="%.4f", key=f"radio_ext_ind_{elemento_id}")
                radio_int = st.number_input(f"Radio Interior (m):", value=default_params.get('radio_int', 0.01), min_value=0.0, max_value=radio_ext*0.99, format="%.4f", key=f"radio_int_ind_{elemento_id}")
                parametros_seccion['radio_ext'] = radio_ext
                parametros_seccion['radio_int'] = radio_int
            elif tipo_seccion == "rectangular":
                lado1 = st.number_input(f"Base (m):", value=default_params.get('lado1', 0.02), min_value=0.001, format="%.4f", key=f"lado1_ind_{elemento_id}")
                lado2 = st.number_input(f"Altura (m):", value=default_params.get('lado2', 0.01), min_value=0.001, format="%.4f", key=f"lado2_ind_{elemento_id}")
                parametros_seccion['lado1'] = lado1
                parametros_seccion['lado2'] = lado2
            elif tipo_seccion == "cuadrada":
                lado = st.number_input(f"Lado (m):", value=default_params.get('lado', 0.02), min_value=0.001, format="%.4f", key=f"lado_ind_{elemento_id}")
                parametros_seccion['lado'] = lado

            # Densidad para análisis dinámico
            if st.session_state.tipo_analisis == "dinamico":
                default_densidad = elemento_existente.get('densidad') or props_material.get('densidad', 2700)
                densidad_sel = st.number_input("Densidad (kg/m³):", value=float(default_densidad), min_value=0.0, format="%.2f", key=f"densidad_ind_{elemento_id}")
            
            # Botón de guardado
            if st.button(f"💾 Guardar Elemento {elemento_id}", key=f"guardar_ind_{elemento_id}"):
                # Recalcular propiedades y matrices
                elem = st.session_state.elementos[i]
                elem['material'] = material_seleccionado
                elem['tipo_seccion'] = tipo_seccion
                elem['parametros_seccion'] = parametros_seccion
                
                area_final = calcular_area_seccion(tipo_seccion, parametros_seccion)
                inercia_final = calcular_momento_inercia(tipo_seccion, parametros_seccion) if st.session_state.tipo_elemento in ["viga", "viga_portico"] else 0
                
                elem['area'] = area_final
                elem['inercia'] = inercia_final
                
                if st.session_state.tipo_analisis == "dinamico":
                    elem['densidad'] = densidad_sel
                
                # Propiedades geométricas (ya deben existir desde el modo interactivo o manual)
                nodo_inicio_obj = next(n for n in st.session_state.nodos if n['id'] == elem['nodo_inicio'])
                nodo_fin_obj = next(n for n in st.session_state.nodos if n['id'] == elem['nodo_fin'])
                
                elem['longitud'] = calcular_longitud_elemento(nodo_inicio_obj, nodo_fin_obj)
                elem['beta'] = calcular_angulo_beta(nodo_inicio_obj, nodo_fin_obj)
                
                E = props_material['modulo_young']
                A, I, L, beta = elem['area'], elem['inercia'], elem['longitud'], elem['beta']
                
                k_global, k_local, m_global, m_local = None, None, None, None
                
                if st.session_state.tipo_elemento == "barra":
                    k_global, k_local = generar_matriz_rigidez_barra(E, A, L, beta)
                    if st.session_state.tipo_analisis == "dinamico":
                        m_local = generar_matriz_masa_barra(elem['densidad'], A, L)
                        m_global = m_local # Placeholder
                elif st.session_state.tipo_elemento == "viga":
                    k_global, k_local = generar_matriz_rigidez_viga(E, I, L)
                    if st.session_state.tipo_analisis == "dinamico":
                        m_local = generar_matriz_masa_viga(elem['densidad'], A, L)
                        m_global = m_local
                else: # viga_portico
                    k_global, k_local = generar_matriz_rigidez_viga_portico(E, A, I, L, beta)
                    if st.session_state.tipo_analisis == "dinamico":
                        m_local = generar_matriz_masa_viga_portico(elem['densidad'], A, L)
                        T = generar_matriz_transformacion_viga_portico(beta)
                        m_global = T.T @ m_local @ T

                st.session_state.matrices_elementos[elemento_id] = {
                    'numerica': k_global.tolist() if k_global is not None else [],
                    'local': k_local.tolist() if k_local is not None else [],
                    'masa_global': m_global.tolist() if m_global is not None else [],
                    'masa_local': m_local.tolist() if m_local is not None else []
                }
                
                st.success(f"✅ Elemento {elemento_id} guardado")
                st.rerun()

    # Verificar si todos los elementos están configurados (tienen material)
    elementos_configurados = all(e.get('material') is not None for e in st.session_state.elementos)

    if st.session_state.elementos:
        st.markdown("### 📋 Elementos Configurados")
        df_elementos = crear_tabla_conectividad()
        st.dataframe(df_elementos, use_container_width=True)
        
        if elementos_configurados:
            if st.button("Continuar →", type="primary"):
                calcular_y_asignar_grados_libertad() # Recalcular GLs por si acaso
                next_step()
        else:
            st.warning("⚠️ Faltan configurar propiedades en algunos elementos. Expanda cada elemento y guárdelo.")
            
elif st.session_state.step == 9:
    
    # --- RAMA: ANÁLISIS ESTÁTICO ---
    if st.session_state.tipo_analisis == "estatico":
        st.markdown("## Configuración de Incógnitas y Conocidos")
        st.markdown("Defina qué grados de libertad son conocidos (desplazamientos/rotaciones) y cuáles son las fuerzas aplicadas.")
        
        if not st.session_state.grados_libertad_info:
            st.warning("Por favor, complete la definición de nodos y elementos primero.")
            st.stop()

        # Crear tabla para entrada de datos
        data_gl = []
        for info in st.session_state.grados_libertad_info:
            # Obtener el nodo y dirección correspondiente
            nodo_obj = next(n for n in st.session_state.nodos if n['id'] == info['nodo'])
            direccion = info['direccion']
            
            # Determinar si es fijo
            es_fijo = nodo_obj['tipo'] == 'fijo'
            
            # Valores por defecto
            despl_conocido = info.get('desplazamiento_conocido', es_fijo) # Si es fijo, el desplazamiento es 0
            val_despl = info.get('valor_desplazamiento', 0.0)
            fuerza_conocida = info.get('fuerza_conocida', not despl_conocido) # Inverso de despl_conocido
            val_fuerza = info.get('valor_fuerza', 0.0)

            data_gl.append({
                'GL': info['numero'],
                'Nodo': info['nodo'],
                'Dirección': direccion.upper(),
                'Fijo': es_fijo,
                'Despl. Conocido': despl_conocido,
                'Valor Despl. [m o rad]': val_despl,
                'Fuerza Conocida': fuerza_conocida,
                'Valor Fuerza [N o Nm]': val_fuerza
            })
        
        df_gl = pd.DataFrame(data_gl)
        
        # Usar st.data_editor para edición interactiva
        edited_df_gl = st.data_editor(
            df_gl,
            column_config={
                "GL": st.column_config.TextColumn("Grado de Libertad", disabled=True),
                "Nodo": st.column_config.TextColumn("Nodo ID", disabled=True),
                "Dirección": st.column_config.TextColumn("Dirección", disabled=True),
                "Fijo": st.column_config.CheckboxColumn("Fijo", disabled=True),
                "Despl. Conocido": st.column_config.CheckboxColumn("Despl. Conocido?"),
                "Valor Despl. [m o rad]": st.column_config.NumberColumn("Valor Desplazamiento", format="%.6f"),
                "Fuerza Conocida": st.column_config.CheckboxColumn("Fuerza Conocida?"),
                "Valor Fuerza [N o Nm]": st.column_config.NumberColumn("Valor Fuerza", format="%.3f")
            },
            use_container_width=True,
            hide_index=True
        )

        if st.button("Continuar →", type="primary"):
            # Actualizar el estado de la sesión con los datos editados
            for index, row in edited_df_gl.iterrows():
                gl_num = int(row['GL'])
                info_gl = next(info for info in st.session_state.grados_libertad_info if info['numero'] == gl_num)
                
                info_gl['desplazamiento_conocido'] = row['Despl. Conocido']
                info_gl['valor_desplazamiento'] = row['Valor Despl. [m o rad]'] if row['Despl. Conocido'] else 0.0
                
                info_gl['fuerza_conocida'] = row['Fuerza Conocida']
                info_gl['valor_fuerza'] = row['Valor Fuerza [N o Nm]'] if row['Fuerza Conocida'] else 0.0
                
                # Lógica de validación
                if row['Despl. Conocido'] and row['Fuerza Conocida']:
                    st.warning(f"GL{gl_num}: No puede conocer la fuerza y el desplazamiento al mismo tiempo. Priorizando desplazamiento.")
                    info_gl['fuerza_conocida'] = False
                    info_gl['valor_fuerza'] = 0.0
                elif not row['Despl. Conocido'] and not row['Fuerza Conocida']:
                    st.warning(f"GL{gl_num}: Debe conocer la fuerza o el desplazamiento. Asumiendo Fuerza = 0.")
                    info_gl['fuerza_conocida'] = True
                    info_gl['valor_fuerza'] = 0.0
            
            next_step()

    # --- RAMA: ANÁLISIS DINÁMICO ---
    elif st.session_state.tipo_analisis == "dinamico":
        st.markdown("## Configuración de Condiciones de Contorno")
        st.markdown("Seleccione los grados de libertad que se encuentran restringidos (desplazamiento nulo).")
        st.info("Los DOF Restringidos (GL fijos) se eliminarán del sistema para el cálculo de autovalores. Solo se calcularán los modos de los DOF Libres.")
        
        if not st.session_state.grados_libertad_info:
            st.warning("Por favor, complete la definición de nodos y elementos primero.")
            st.stop()

        st.markdown("### Grados de Libertad Restringidos")
        
        # Crear matriz de selección
        cols = st.columns(3)
        col_idx = 0
        
        # Obtener información de nodos fijos
        nodos_fijos_ids = {n['id'] for n in st.session_state.nodos if n.get('tipo', 'libre') == 'fijo'}

        for i, info in enumerate(st.session_state.grados_libertad_info):
            with cols[col_idx % 3]:
                nodo = info['nodo']
                direccion = info['direccion']
                gl_num = info['numero']
                
                # Determinar si está restringido por defecto (nodo 'fijo')
                es_fijo_por_defecto = (nodo in nodos_fijos_ids)
                
                # Checkbox para marcar como restringido
                es_restringido = st.checkbox(
                    f"GL{gl_num}: Nodo {nodo} - {direccion}",
                    value=st.session_state.condiciones_contorno_dinamica.get(gl_num, es_fijo_por_defecto), # Valor por defecto
                    key=f"restriccion_{gl_num}",
                    help="Este GL está restringido porque el nodo es 'Fijo'." if es_fijo_por_defecto else "Marque para restringir este GL.",
                    disabled=es_fijo_por_defecto # <-- ESTA ES LA LÍNEA CLAVE
                )
                
                # Asegurar que el estado refleje la restricción
                if es_fijo_por_defecto:
                    st.session_state.condiciones_contorno_dinamica[gl_num] = True
                else:
                    st.session_state.condiciones_contorno_dinamica[gl_num] = es_restringido
            
            col_idx += 1
        
        st.divider()
        
        # Resumen de restricciones
        dof_restringidos_list = [gl for gl, v in st.session_state.condiciones_contorno_dinamica.items() if v]
        dof_libres = [info['numero'] for info in st.session_state.grados_libertad_info if not st.session_state.condiciones_contorno_dinamica.get(info['numero'], False)]
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("DOF Restringidos", len(dof_restringidos_list))
            if dof_restringidos_list:
                st.write(f"GL: {', '.join([str(gl) for gl in sorted(dof_restringidos_list)])}")
        
        with col2:
            st.metric("DOF Libres (Modos a calcular)", len(dof_libres))
            if dof_libres:
                st.write(f"GL: {', '.join([str(gl) for gl in sorted(dof_libres)])}")
        
        if len(dof_libres) < 1:
            st.error("Debe tener al menos 1 grado de libertad libre para realizar el análisis dinámico")
        else:
            st.success(f"Sistema válido: Se calcularán {len(dof_libres)} modos de vibración.")
        
        st.divider()
        
        if st.button("Continuar →", type="primary"):
            next_step()

# --- PASO 10 (UNIFICADO) ---
# --- PASO 10 (UNIFICADO) ---
elif st.session_state.step == 10:
    
    # --- RAMA: ANÁLISIS ESTÁTICO ---
    if st.session_state.tipo_analisis == "estatico":
        st.markdown("## Resultados del Análisis Estático")
        
        # Mostrar botón de cálculo si no hay resultados
        if not st.session_state.resultados:
            if st.button("🧮 Calcular Sistema Estático", type="primary", use_container_width=True):
                if not st.session_state.elementos or not st.session_state.grados_libertad_info:
                    st.error("Por favor, asegúrese de haber definido todos los nodos y elementos.")
                else:
                    resultado_estatico = resolver_sistema()
                    if resultado_estatico and resultado_estatico.get('exito'):
                        st.session_state.resultados = resultado_estatico
                        st.success("Análisis estático completado exitosamente.")
                        st.rerun() # Recargar para mostrar resultados
                    else:
                        st.error("Error al resolver el sistema estático. Verifique las condiciones de contorno.")
            
            st.info("Presione 'Calcular Sistema Estático' para resolver la estructura.")

        # Mostrar resultados si existen
        if st.session_state.resultados and st.session_state.resultados.get('exito'):
            resultado_estatico = st.session_state.resultados
            
            st.markdown("### Métricas Principales")
            col1, col2, col3, col4 = st.columns(4)
            with col1: st.metric("Nodos", len(st.session_state.nodos))
            with col2: st.metric("Elementos", len(st.session_state.elementos))
            with col3: st.metric("DOF Libres", len(st.session_state.grados_libertad_info))
            with col4: st.metric("Determinante K (Global)", f"{resultado_estatico['determinante']:.3e}")
            
            st.divider()
            
            st.markdown("### Desplazamientos Nodales")
            df_desplazamientos = pd.DataFrame({
                'GL': [info['numero'] for info in st.session_state.grados_libertad_info],
                'Nodo': [info['nodo'] for info in st.session_state.grados_libertad_info],
                'Dirección': [info['direccion'] for info in st.session_state.grados_libertad_info],
                'Desplazamiento [m o rad]': [formatear_unidades(d, "desplazamiento") for d in resultado_estatico['desplazamientos']]
            })
            st.dataframe(df_desplazamientos, use_container_width=True, hide_index=True)
            
            st.divider()
            
            st.markdown("### Fuerzas y Reacciones")
            df_fuerzas = pd.DataFrame({
                'GL': [info['numero'] for info in st.session_state.grados_libertad_info],
                'Nodo': [info['nodo'] for info in st.session_state.grados_libertad_info],
                'Dirección': [info['direccion'] for info in st.session_state.grados_libertad_info],
                'Fuerza [N o Nm]': [formatear_unidades(f, "fuerza") for f in resultado_estatico['fuerzas']],
                'Tipo': ["Aplicada" if info['fuerza_conocida'] else "Reacción" for info in st.session_state.grados_libertad_info]
            })
            st.dataframe(df_fuerzas, use_container_width=True, hide_index=True)
            
            st.divider()

            st.markdown("### 📊 Visualización de la Estructura")
            factor_escala_on_screen = st.slider("Factor de escala para visualización:", 1, 1000, 100, key="factor_escala_pantalla")
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("#### Estructura Deformada")
                fig_deformada = visualizar_estructura_moderna(mostrar_deformada=True, factor_escala=factor_escala_on_screen)
                if fig_deformada:
                    # --- CORRECCIÓN: Quitar use_container_width ---
                    st.pyplot(fig_deformada)
            with col2:
                st.markdown("#### Estructura Original")
                fig_original = visualizar_estructura_moderna(mostrar_deformada=False)
                if fig_original:
                    # --- CORRECCIÓN: Quitar use_container_width ---
                    st.pyplot(fig_original)
            st.divider()
            
            st.markdown("### Exportar Análisis Estático")
            col_exp1, col_exp2 = st.columns(2)
            with col_exp1:
                pdf_reporte_estatico = generar_pdf_reporte_estatico()
                if pdf_reporte_estatico:
                    st.download_button(label="Descargar Reporte PDF", data=pdf_reporte_estatico, file_name=f"analisis_estatico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf", mime="application/pdf", type="primary", use_container_width=True)
            
            with col_exp2:
                if OPENPYXL_AVAILABLE:
                    excel_reporte_estatico = generar_excel_reporte_estatico()
                    if excel_reporte_estatico:
                        st.download_button(label="Descargar Reporte Excel", data=excel_reporte_estatico, file_name=f"analisis_estatico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
                else:
                    st.warning("Instale 'openpyxl' para exportar a Excel.")
            
            st.divider()
            
            with st.expander("Ver Tablas de Referencia (Nodos, Elementos, K Global)"):
                st.markdown("#### Tabla de Nodos")
                df_nodos = crear_tabla_nodos()
                st.dataframe(df_nodos, use_container_width=True, hide_index=True)
                
                st.markdown("#### Tabla de Conectividad")
                df_conectividad = crear_tabla_conectividad()
                st.dataframe(df_conectividad, use_container_width=True, hide_index=True)
                
                st.markdown("#### Matriz de Rigidez Global (K)")
                K_global = resultado_estatico['K_global']
                df_K_global = pd.DataFrame(K_global, 
                                           columns=[f"GL{i+1}" for i in range(K_global.shape[1])],
                                           index=[f"GL{i+1}" for i in range(K_global.shape[0])])
                st.dataframe(df_K_global.applymap(lambda x: f"{x:.3e}"), use_container_width=True)

    # --- RAMA: ANÁLISIS DINÁMICO ---
    elif st.session_state.tipo_analisis == "dinamico":
        st.markdown("## Resultados del Análisis Dinámico")
        st.markdown("Frecuencias naturales y modos de vibración")
        
        # Mostrar botón de cálculo si no hay resultados
        if not st.session_state.resultados_dinamicos:
            if st.button("🧮 Calcular Sistema Dinámico", type="primary", use_container_width=True):
                if not st.session_state.elementos or not st.session_state.grados_libertad_info:
                    st.error("Por favor, asegúrese de haber definido todos los nodos y elementos.")
                elif len([gl for gl in st.session_state.grados_libertad_info if not st.session_state.condiciones_contorno_dinamica.get(gl['numero'], False)]) < 1:
                    st.error("Debe tener al menos 1 grado de libertad libre para realizar el análisis dinámico.")
                else:
                    resultado = resolver_sistema_dinamico()
                    if resultado and resultado.get('exito'):
                        st.session_state.resultados_dinamicos = resultado
                        st.success("Sistema dinámico resuelto exitosamente.")
                        st.rerun() # Recargar para mostrar resultados
                    else:
                        st.error("Error al resolver el sistema dinámico. Verifique las matrices o las condiciones de contorno.")
            
            st.warning("Por favor, presione 'Calcular Sistema Dinámico' para ver los resultados.")

        # Mostrar resultados si existen
        if st.session_state.resultados_dinamicos and st.session_state.resultados_dinamicos.get('exito'):
            resultado_din = st.session_state.resultados_dinamicos
            
            # Métricas principales
            st.markdown("### Métricas Principales")
            col1, col2, col3, col4 = st.columns(4)
            with col1: st.metric("Nodos", len(st.session_state.nodos))
            with col2: st.metric("Elementos", len(st.session_state.elementos))
            with col3: st.metric("DOF Libres", len(resultado_din['dof_libres']))
            with col4: st.metric("Modos Calculados", len(resultado_din['frecuencias_hz']))
            
            st.divider()
            
            # Tabla de todos los modos (Formato Excel)
            st.markdown("### Frecuencias y Formas Modales")
            st.info("Tabla de resultados estilo Excel, como la solicitada.")
            
            df_modos_completa = crear_tabla_modos_completa()
            st.dataframe(df_modos_completa, use_container_width=True, hide_index=True)
            
            st.divider()
            
            # Selector de modo para visualización
            st.markdown("### Visualización de Modos")
            num_modos = len(resultado_din['frecuencias_hz'])

            if num_modos > 0:
                modo_seleccionado = st.selectbox(
                    "Seleccione el modo a visualizar",
                    range(1, num_modos + 1),
                    key="modo_visualizacion_selector",
                    format_func=lambda x: f"Modo {x} - f = {resultado_din['frecuencias_hz'][x-1]:.2f} Hz"
                )
                
                st.session_state.modo_visualizacion = modo_seleccionado
                
                idx_modo = modo_seleccionado - 1 
                omega_modo = resultado_din['frecuencias_rad'][idx_modo]
                f_modo = resultado_din['frecuencias_hz'][idx_modo]
                
                col_m1, col_m2, col_m3 = st.columns(3)
                with col_m1: st.metric("Número de Modo", modo_seleccionado)
                with col_m2: st.metric("Frecuencia (f)", f"{f_modo:.4f} Hz")
                with col_m3: st.metric("Frec. Angular (ω)", f"{omega_modo:.4f} rad/s")

                # Calcular el factor de escala automático
                factor_escala = None 
                try:
                    max_despl = np.max(np.abs(resultado_din['eigenvectors'][:, idx_modo]))
                    if max_despl > 1e-9:
                        rango_x_orig = max(n['x'] for n in st.session_state.nodos) - min(n['x'] for n in st.session_state.nodos)
                        rango_y_orig = max(n['y'] for n in st.session_state.nodos) - min(n['y'] for n in st.session_state.nodos)
                        rango_global_orig = max(rango_x_orig, rango_y_orig, 1.0)
                        factor_escala = (0.1 * rango_global_orig) / max_despl
                        factor_escala = max(1, min(factor_escala, 500))
                    else:
                        factor_escala = 1.0
                except:
                    factor_escala = 1.0 
                
                
                col_viz1, col_viz2 = st.columns(2) 
                
                with col_viz1:
                    st.markdown("#### Estructura Original")
                    fig_original_base = visualizar_estructura_moderna(mostrar_deformada=False, factor_escala=1.0)
                    if fig_original_base:
                        # --- CORRECCIÓN: Quitar use_container_width ---
                        st.pyplot(fig_original_base)
                    else:
                        st.warning("No se pudo generar el gráfico de la estructura original.")

                with col_viz2:
                    st.markdown(f"#### Modo {modo_seleccionado} (f = {f_modo:.2f} Hz)")
                    fig_modo = visualizar_modo_dinamico(idx_modo, factor_escala=factor_escala) 
                    if fig_modo:
                        # --- CORRECCIÓN: Quitar use_container_width ---
                        st.pyplot(fig_modo) 
                    else:
                        st.warning("No se pudo generar el gráfico del modo de vibración.")
                
            else:
                st.warning("⚠️ No se encontraron modos de vibración (frecuencias > 0).")
                st.info("Esto usualmente significa que la estructura no está correctamente restringida (es un mecanismo) y solo tiene modos de cuerpo rígido con frecuencia cero.")

            st.divider()
            
            st.markdown("### Exportar Análisis Dinámico")
            col_exp1, col_exp2 = st.columns(2)
            
            with col_exp1:
                pdf_reporte = generar_pdf_reporte_dinamico()
                if pdf_reporte:
                    st.download_button(label="Descargar Reporte PDF (Tablas y Gráficos)", data=pdf_reporte, file_name=f"analisis_dinamico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf", mime="application/pdf", type="primary", use_container_width=True)
            with col_exp2:
                if OPENPYXL_AVAILABLE:
                    excel_reporte = generar_excel_reporte_dinamico()
                    if excel_reporte:
                        st.download_button(label="Descargar Reporte Excel (Estilo Solicitado)", data=excel_reporte, file_name=f"analisis_dinamico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
                else:
                    st.warning("Instale 'openpyxl' para exportar a Excel.")

            st.divider()
            
            st.markdown("#### Tabla de Nodos")
            df_nodos = crear_tabla_nodos()
            st.dataframe(df_nodos, use_container_width=True, hide_index=True)
            st.markdown("#### Tabla de Conectividad")
            df_conectividad = crear_tabla_conectividad()
            st.dataframe(df_conectividad, use_container_width=True, hide_index=True)

            with st.expander("Ver Matrices Globales del Sistema (K y M)"):
                st.markdown(f"**Matrices Globales (Tamaño: {len(st.session_state.grados_libertad_info)} x {len(st.session_state.grados_libertad_info)})**")
                K_global_full = resultado_din['K_global']
                M_global_full = resultado_din['M_global']
                gl_labels = [f"GL{info['numero']}" for info in st.session_state.grados_libertad_info]
                
                st.markdown("#### Matriz de Rigidez Global (K)")
                df_K_global = pd.DataFrame(K_global_full, columns=gl_labels, index=gl_labels)
                st.dataframe(df_K_global.applymap(lambda x: f"{x:.3e}"), use_container_width=True)

                st.markdown("#### Matriz de Masa Global (M)")
                df_M_global = pd.DataFrame(M_global_full, columns=gl_labels, index=gl_labels)
                st.dataframe(df_M_global.applymap(lambda x: f"{x:.3e}"), use_container_width=True)
                
                st.markdown("---")
                st.markdown(f"**Matrices 'Libres' Usadas en el Solver (Tamaño: {len(resultado_din['dof_libres'])} x {len(resultado_din['dof_libres'])})**")
                gl_labels_libres = [f"GL{gl_num}" for gl_num in resultado_din['dof_libres']]

                st.markdown("#### Matriz de Rigidez Libre (K_libre)")
                df_K_libre = pd.DataFrame(resultado_din['K_libre'], columns=gl_labels_libres, index=gl_labels_libres)
                st.dataframe(df_K_libre.applymap(lambda x: f"{x:.3e}"), use_container_width=True)
                
                st.markdown("#### Matriz de Masa Libre (M_libre)")
                df_M_libre = pd.DataFrame(resultado_din['M_libre'], columns=gl_labels_libres, index=gl_labels_libres)
                st.dataframe(df_M_libre.applymap(lambda x: f"{x:.3e}"), use_container_width=True)

            with st.expander("Ver Matrices Locales de Elementos (K' y M')"):
                if st.session_state.tipo_elemento == "viga_portico":
                    labels = ["u1'", "v1'", "θ1'", "u2'", "v2'", "θ2'"]
                elif st.session_state.tipo_elemento == "viga":
                    labels = ["v1'", "θ1'", "v2'", "θ2'"]
                else:
                    labels = ["u1'", "v1'", "u2'", "v2'"]

                for elemento in st.session_state.elementos:
                    st.markdown(f"#### Elemento {elemento['id']} (Nodos {elemento['nodo_inicio']} → {elemento['nodo_fin']})")
                    col_k, col_m = st.columns(2)
                    
                    with col_k:
                        st.markdown("**Matriz de Rigidez Local (k')**")
                        matriz_k_local = np.array(st.session_state.matrices_elementos[elemento['id']].get('local', []))
                        if matriz_k_local.any():
                            df_k_local = pd.DataFrame(matriz_k_local, index=labels, columns=labels)
                            st.dataframe(df_k_local.applymap(lambda x: f"{x:.3e}"))
                    
                    with col_m:
                        st.markdown("**Matriz de Masa Local (m')**")
                        matriz_m_local = np.array(st.session_state.matrices_elementos[elemento['id']].get('masa_local', []))
                        if matriz_m_local.any():
                            df_m_local = pd.DataFrame(matriz_m_local, index=labels, columns=labels)
                            st.dataframe(df_m_local.applymap(lambda x: f"{x:.3e}"))

# El paso 11 ya no se usa
elif st.session_state.step == 11:
    st.markdown("## Resultados del Análisis")
    st.info("Los resultados se muestran en el paso anterior (Paso 10).")
    st.warning("Por favor, retroceda al paso anterior para ver los resultados.")
    if st.button("← Volver a Resultados"):
        st.session_state.step = 10
        st.rerun()

# --- Footer (de V4.7) ---
st.markdown("""
<div style='background-color: #212529; padding: 2rem; margin-top: 3rem; border-radius: 15px;'>
    <div style='text-align: center; color: white;'>
        <h3 style='color: white; margin-bottom: 1rem;'>📝 Sistema de Análisis Estructural Avanzado</h3>
        <p style='color: rgba(255,255,255,0.8); line-height: 1.6;'>
            Desarrollado con soporte para múltiples tipos de elementos estructurales:<br>
            <strong>Barras</strong> • <strong>Vigas</strong> • <strong>Vigas Pórtico</strong><br>
            Incluye análisis Estático y Dinámico (Modos y Frecuencias).<br>
            <strong>Modo Manual</strong> e <strong>Interactivo</strong> disponibles para todos los tipos de elementos.
        </p>
    </div>
</div>
""", unsafe_allow_html=True)